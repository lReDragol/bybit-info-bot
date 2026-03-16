import os
import json
import time
import logging
import threading
import sqlite3
import shutil
import statistics
import atexit
import functools
import math
import re
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from http.cookies import SimpleCookie
from time import sleep
from urllib.parse import parse_qs, urlparse

import requests
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
from matplotlib.ticker import MaxNLocator
from openpyxl import Workbook, load_workbook
import telebot
from telebot import types
from telebot.apihelper import ApiTelegramException


# ------------------ НАСТРОЙКИ И КОНФИГУРАЦИЯ ------------------

logging.basicConfig(level=logging.ERROR)
plt.switch_backend('Agg')

MESSAGES = {
    'start_message': 'Бот запущен!',
    'menu_balance': 'Баланс',
    'menu_graph': 'График',
    'menu_top': 'Топ',
    'menu_admin': 'Админ',
    'error_balance': 'Ошибка получения баланса',
    'error_graph': 'Ошибка генерации графика',
    'admin_no_access': 'У вас нет прав доступа.',
    'migrate_ok': 'Миграция прошла успешно.',
    'migrate_fail': 'Ошибка миграции.',
    'gen_images_done': 'Генерация картинок завершена.',
    'admin_panel_title': 'Панель админа',
    'admin_download_not_found': 'Файл базы данных не найден.',
    'config_title': 'Конфигурация',
    'admin_reload_success': 'Конфиг перечитан и применён без перезапуска процесса.',
    'admin_resume_bot': 'Бот обновился.'
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
CACHE_DIR = os.path.join(BASE_DIR, "cache")
GRAPH_DIR = os.path.join(CACHE_DIR, "graphs")
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
INSTANCE_LOCK_FILE = os.path.join(CACHE_DIR, "tgbybit.lock")
TOP_BOTS_IMAGE_FILE = os.path.join(CACHE_DIR, "top_bots.png")
TOP_BOTS_PAGE_SIZE = 8
TOP_SORT_MODES = {
    "earnings": {
        "button": "Заработок",
        "title": "общему заработку",
        "caption_label": "Итог",
        "overview_label": "Итог"
    },
    "pnl": {
        "button": "P&L",
        "title": "P&L",
        "caption_label": "P&L",
        "overview_label": "P&L"
    },
    "percent": {
        "button": "% ROI",
        "title": "доходности %",
        "caption_label": "% ROI",
        "overview_label": "% ROI"
    }
}
NOTIFICATION_SETTINGS_DEFAULTS = {
    "bot_liquidation": True,
    "bot_stop_loss": True,
    "bot_trailing_stop": False,
    "bot_manual_close": False,
    "market_drop": True,
    "bot_risk_limit": True,
    "bot_pnl_drawdown": True,
    "bot_liq_distance": True,
    "bot_repeat_loss": True
}
NOTIFICATION_LABELS = {
    "bot_liquidation": "Ликвидация",
    "bot_stop_loss": "Стоп-лосс",
    "bot_trailing_stop": "Трейлинг-выход",
    "bot_manual_close": "Ручное закрытие",
    "market_drop": "Падение рынка",
    "bot_risk_limit": "Риск-лимиты",
    "bot_pnl_drawdown": "Просадка P&L",
    "bot_liq_distance": "Близко к ликвидации",
    "bot_repeat_loss": "Повтор после лося"
}
RISK_SETTINGS_DEFAULTS = {
    "max_total_margin_per_symbol_usdt": 250.0,
    "max_active_bots_same_symbol_direction": 1.0,
    "max_leverage_grid_futures": 10.0,
    "max_leverage_mart_futures": 15.0,
    "max_loss_alert_pct": -15.0,
    "min_liq_distance_pct": 8.0,
    "margin_growth_alert_pct": 25.0,
    "repeat_loss_cooldown_hours": 24.0
}
API_SETTINGS_DEFAULTS = {
    "enabled": True,
    "host": "127.0.0.1",
    "port": 8877,
    "token": ""
}


def apply_config_defaults(config_data):
    normalized = dict(config_data or {})
    raw_settings = normalized.get("notification_settings") or {}
    merged_settings = dict(NOTIFICATION_SETTINGS_DEFAULTS)
    for key in NOTIFICATION_SETTINGS_DEFAULTS:
        if key in raw_settings:
            merged_settings[key] = bool(raw_settings[key])
    normalized["notification_settings"] = merged_settings
    raw_risk_settings = normalized.get("risk_settings") or {}
    merged_risk_settings = dict(RISK_SETTINGS_DEFAULTS)
    for key, default_value in RISK_SETTINGS_DEFAULTS.items():
        if key not in raw_risk_settings:
            continue
        try:
            merged_risk_settings[key] = float(raw_risk_settings[key])
        except (TypeError, ValueError):
            merged_risk_settings[key] = default_value
    normalized["risk_settings"] = merged_risk_settings
    raw_api_settings = normalized.get("api_settings") or {}
    merged_api_settings = dict(API_SETTINGS_DEFAULTS)
    for key, default_value in API_SETTINGS_DEFAULTS.items():
        if key not in raw_api_settings:
            continue
        if isinstance(default_value, bool):
            merged_api_settings[key] = bool(raw_api_settings[key])
        elif isinstance(default_value, int):
            try:
                merged_api_settings[key] = int(raw_api_settings[key])
            except (TypeError, ValueError):
                merged_api_settings[key] = default_value
        else:
            merged_api_settings[key] = str(raw_api_settings[key])
    normalized["api_settings"] = merged_api_settings
    if "bot_close_notify_bootstrapped" not in normalized:
        normalized["bot_close_notify_bootstrapped"] = False
    return normalized


def load_config():
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return apply_config_defaults(json.load(f))


def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)


if os.name == "nt":
    import msvcrt
else:
    import fcntl


INSTANCE_LOCK_HANDLE = None


def acquire_instance_lock():
    global INSTANCE_LOCK_HANDLE
    os.makedirs(CACHE_DIR, exist_ok=True)
    lock_handle = open(INSTANCE_LOCK_FILE, 'a+', encoding='utf-8')
    try:
        lock_handle.seek(0)
        if os.name == "nt":
            msvcrt.locking(lock_handle.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError:
        lock_handle.close()
        return False, ""

    lock_handle.seek(0)
    lock_handle.truncate()
    lock_handle.write(str(os.getpid()))
    lock_handle.flush()
    INSTANCE_LOCK_HANDLE = lock_handle
    return True, str(os.getpid())


def release_instance_lock():
    global INSTANCE_LOCK_HANDLE
    if INSTANCE_LOCK_HANDLE is None:
        return
    try:
        INSTANCE_LOCK_HANDLE.seek(0)
        if os.name == "nt":
            msvcrt.locking(INSTANCE_LOCK_HANDLE.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            fcntl.flock(INSTANCE_LOCK_HANDLE.fileno(), fcntl.LOCK_UN)
    except OSError:
        pass
    finally:
        INSTANCE_LOCK_HANDLE.close()
        INSTANCE_LOCK_HANDLE = None


atexit.register(release_instance_lock)


def handler_guard(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except requests.RequestException as e:
            logging.error("Telegram request error in %s: %s", func.__name__, e)
            return None
        except ApiTelegramException as e:
            logging.error("Telegram API error in %s: %s", func.__name__, e)
            return None
        except Exception:
            logging.exception("Handler error in %s", func.__name__)
            return None
    return wrapper


config = load_config()
TOKEN = config.get('TOKEN', '')
cookies = config.get('cookies', '')
admins = config.get('admins', [])
db_update_interval = config.get('db_update_interval', 30)
balance_send_interval = config.get('balance_send_interval', 30)
chat_id = config.get('chat_id', '')


def get_notification_settings():
    settings = dict(NOTIFICATION_SETTINGS_DEFAULTS)
    settings.update(config.get("notification_settings") or {})
    return settings


def get_risk_settings():
    settings = dict(RISK_SETTINGS_DEFAULTS)
    settings.update(config.get("risk_settings") or {})
    return settings


def get_api_settings():
    settings = dict(API_SETTINGS_DEFAULTS)
    settings.update(config.get("api_settings") or {})
    return settings

REQUEST_TIMEOUT = 60
MAX_RETRIES = 5
EXCEL_FILE = os.path.join(BASE_DIR, "balance_data.xlsx")
DB_FILE = os.path.join(BASE_DIR, "balance_data.db")
WAITING_FOR_RENEW = False
BOT_PAGE_SIZE = 50
BOT_HISTORY_STATUS = 1
BOT_ARCHIVE_SYNC_INTERVAL_MINUTES = 15
BOT_DUPLICATE_MATCH_ABS_USDT = 5.0
BOT_DUPLICATE_MATCH_RATIO = 0.03
BOT_DUPLICATE_MIN_JUMP_USDT = 30.0
RUB_CACHE = {
    "value": None,
    "updated_ts": 0.0,
    "ttl_seconds": 600,
    "stale_ttl_seconds": 86400
}
GRAPH_CACHE_STATE = {
    "last_cleanup_ts": 0.0
}
BOT_SNAPSHOT_TABLE_SQL = '''
    CREATE TABLE IF NOT EXISTS bot_snapshots (
        snapshot_time TEXT NOT NULL,
        bot_index INTEGER NOT NULL,
        bot_id TEXT,
        symbol TEXT,
        bot_type TEXT,
        title TEXT,
        badge TEXT,
        investment_usdt REAL,
        pnl_usdt REAL,
        equity_usdt REAL,
        pnl_percent REAL,
        status TEXT,
        display_status TEXT,
        is_active INTEGER,
        close_code TEXT,
        PRIMARY KEY (snapshot_time, bot_index)
    )
'''
BOT_SNAPSHOT_INDEX_SQL = '''
    CREATE INDEX IF NOT EXISTS idx_bot_snapshots_lookup
    ON bot_snapshots(symbol, bot_type, snapshot_time)
'''
BOT_SNAPSHOT_ID_INDEX_SQL = '''
    CREATE INDEX IF NOT EXISTS idx_bot_snapshots_bot_id
    ON bot_snapshots(bot_id, snapshot_time)
'''
BOT_ARCHIVE_TABLE_SQL = '''
    CREATE TABLE IF NOT EXISTS bot_archive (
        bot_id TEXT PRIMARY KEY,
        symbol TEXT,
        bot_type TEXT,
        title TEXT,
        badge TEXT,
        status TEXT,
        display_status TEXT,
        close_code TEXT,
        close_reason TEXT,
        investment_usdt REAL,
        pnl_usdt REAL,
        equity_usdt REAL,
        pnl_percent REAL,
        final_profit_usdt REAL,
        settlement_assets_text TEXT,
        settlement_assets_usdt REAL,
        leverage TEXT,
        mode TEXT,
        created_ts INTEGER,
        ended_ts INTEGER,
        first_seen_at TEXT,
        last_seen_at TEXT,
        last_snapshot_time TEXT,
        is_active INTEGER,
        raw_json TEXT,
        close_notified_at TEXT,
        close_notify_type TEXT
    )
'''
BOT_ARCHIVE_TOP_INDEX_SQL = '''
    CREATE INDEX IF NOT EXISTS idx_bot_archive_top
    ON bot_archive(final_profit_usdt, pnl_usdt, is_active)
'''
BOT_ARCHIVE_ID_INDEX_SQL = '''
    CREATE UNIQUE INDEX IF NOT EXISTS idx_bot_archive_bot_id
    ON bot_archive(bot_id)
'''
ALERT_EVENTS_TABLE_SQL = '''
    CREATE TABLE IF NOT EXISTS alert_events (
        alert_key TEXT PRIMARY KEY,
        alert_type TEXT,
        bot_id TEXT,
        symbol TEXT,
        created_at TEXT,
        payload_json TEXT
    )
'''
ALERT_EVENTS_INDEX_SQL = '''
    CREATE INDEX IF NOT EXISTS idx_alert_events_lookup
    ON alert_events(alert_type, created_at)
'''

# Глобальная переменная для остановки потоков
stop_threads = False
thread_run_token = 0
SCHEDULER_LOCK = threading.Lock()
SCHEDULER_STATE = {
    "db_slot": None,
    "balance_slot": None,
    "bot_archive_slot": None
}
API_SERVER_STATE = {
    "server": None,
    "thread": None
}
EFFECTIVE_BALANCE_SQL = (
    "CASE WHEN balance_in_usd IS NOT NULL AND balance_in_usd > 0 "
    "THEN balance_in_usd ELSE current_balance END"
)
BALANCE_REPAIR_DUPLICATE_SECONDS = 45
BALANCE_REPAIR_MAX_DUPLICATE_DIFF_PCT = 3.0
BALANCE_REPAIR_MAX_SEGMENT_ROWS = 8
BALANCE_REPAIR_MAX_SPAN_MINUTES = 20
BALANCE_REPAIR_MAX_BOUNDARY_GAP_PCT = 5.0
BALANCE_REPAIR_MIN_DROP_PCT = 12.0
BALANCE_REPAIR_MIN_DROP_USDT = 15.0
BALANCE_REPAIR_SEGMENT_CEILING_RATIO = 0.92

# Если БД существует, будем использовать её
USE_DB = os.path.exists(DB_FILE)

bot = telebot.TeleBot(TOKEN)

user_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
user_keyboard.add(
    types.KeyboardButton(MESSAGES['menu_balance']),
    types.KeyboardButton(MESSAGES['menu_graph'])
)
user_keyboard.add(
    types.KeyboardButton(MESSAGES['menu_top']),
    types.KeyboardButton(MESSAGES['menu_admin'])
)


# ------------------ РАБОТА С EXCEL И/ИЛИ БД ------------------

def setup_excel():
    try:
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        # Для Excel сохраняются только базовые 4 поля
        worksheet.append(['Дата', 'current_balance', 'balance_rub', 'change_percent'])
        workbook.save(EXCEL_FILE)
    return workbook, worksheet


if not USE_DB:
    workbook, worksheet = setup_excel()


# --- Работа с SQLite ---
def get_db_connection():
    return sqlite3.connect(DB_FILE)


def create_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
       CREATE TABLE IF NOT EXISTS balances (
         date TEXT PRIMARY KEY,
         current_balance REAL,
         balance_rub REAL,
         change_percent REAL,
         balance_in_usd REAL,
         balance_in_btc REAL,
         profit_in_usd REAL,
         profit_in_btc REAL,
         pnl_percentage REAL,
         current_profit_in_usd REAL,
         current_profit_in_btc REAL,
         current_pnl_percentage REAL,
         origin_balance REAL,
         bot_balance REAL,
         funding_balance REAL,
         non_bot_balance REAL,
         update_interval INTEGER
       )
    ''')
    cursor.execute(BOT_SNAPSHOT_TABLE_SQL)
    cursor.execute(BOT_ARCHIVE_TABLE_SQL)
    cursor.execute(ALERT_EVENTS_TABLE_SQL)
    conn.commit()
    conn.close()
    ensure_db_schema()


def ensure_table_columns(cursor, table_name, expected_columns):
    cursor.execute(f"PRAGMA table_info({table_name})")
    existing_columns = {row[1] for row in cursor.fetchall()}
    for column_name, column_type in expected_columns.items():
        if column_name in existing_columns:
            continue
        try:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")
        except Exception as e:
            logging.error(f"Ошибка при добавлении столбца {column_name} в {table_name}: {e}")


def ensure_db_schema():
    balance_expected = {
        'current_balance': 'REAL',
        'balance_rub': 'REAL',
        'change_percent': 'REAL',
        'balance_in_usd': 'REAL',
        'balance_in_btc': 'REAL',
        'profit_in_usd': 'REAL',
        'profit_in_btc': 'REAL',
        'pnl_percentage': 'REAL',
        'current_profit_in_usd': 'REAL',
        'current_profit_in_btc': 'REAL',
        'current_pnl_percentage': 'REAL',
        'origin_balance': 'REAL',
        'bot_balance': 'REAL',
        'funding_balance': 'REAL',
        'non_bot_balance': 'REAL',
        'update_interval': 'INTEGER'
    }
    bot_snapshot_expected = {
        'bot_id': 'TEXT',
        'status': 'TEXT',
        'display_status': 'TEXT',
        'is_active': 'INTEGER',
        'close_code': 'TEXT'
    }
    bot_archive_expected = {
        'bot_id': 'TEXT',
        'symbol': 'TEXT',
        'bot_type': 'TEXT',
        'title': 'TEXT',
        'badge': 'TEXT',
        'status': 'TEXT',
        'display_status': 'TEXT',
        'close_code': 'TEXT',
        'close_reason': 'TEXT',
        'investment_usdt': 'REAL',
        'pnl_usdt': 'REAL',
        'equity_usdt': 'REAL',
        'pnl_percent': 'REAL',
        'final_profit_usdt': 'REAL',
        'settlement_assets_text': 'TEXT',
        'settlement_assets_usdt': 'REAL',
        'leverage': 'TEXT',
        'mode': 'TEXT',
        'created_ts': 'INTEGER',
        'ended_ts': 'INTEGER',
        'first_seen_at': 'TEXT',
        'last_seen_at': 'TEXT',
        'last_snapshot_time': 'TEXT',
        'is_active': 'INTEGER',
        'raw_json': 'TEXT',
        'close_notified_at': 'TEXT',
        'close_notify_type': 'TEXT'
    }
    conn = get_db_connection()
    cursor = conn.cursor()
    ensure_table_columns(cursor, 'balances', balance_expected)
    cursor.execute(BOT_SNAPSHOT_TABLE_SQL)
    ensure_table_columns(cursor, 'bot_snapshots', bot_snapshot_expected)
    cursor.execute(BOT_SNAPSHOT_INDEX_SQL)
    cursor.execute(BOT_SNAPSHOT_ID_INDEX_SQL)
    cursor.execute(BOT_ARCHIVE_TABLE_SQL)
    ensure_table_columns(cursor, 'bot_archive', bot_archive_expected)
    cursor.execute(BOT_ARCHIVE_ID_INDEX_SQL)
    cursor.execute(BOT_ARCHIVE_TOP_INDEX_SQL)
    cursor.execute(ALERT_EVENTS_TABLE_SQL)
    cursor.execute(ALERT_EVENTS_INDEX_SQL)
    conn.commit()
    conn.close()


# Миграция данных из Excel в БД (если требуется)
def migrate_excel_to_db():
    try:
        create_db()
        ensure_db_schema()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM balances")
        wb, ws = setup_excel()
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Дата":
                continue
            date_val = row[0]
            current_balance = row[1]
            balance_rub = row[2]
            change_percent = row[3]
            # Остальные поля запишем как 0
            cursor.execute(
                "INSERT OR REPLACE INTO balances (date, current_balance, balance_rub, change_percent, balance_in_usd, balance_in_btc, profit_in_usd, profit_in_btc, pnl_percentage, current_profit_in_usd, current_profit_in_btc, current_pnl_percentage, origin_balance, bot_balance, funding_balance, non_bot_balance, update_interval) VALUES (?, ?, ?, ?, 0, 0, 0, 0, 0, 0, 0, 0, ?, ?, 0, 0, ?)",
                (date_val, current_balance, balance_rub, change_percent, current_balance, current_balance,
                 config.get('db_update_interval', 30))
            )
        conn.commit()
        conn.close()
        global USE_DB
        USE_DB = True
        return True
    except Exception as e:
        logging.error(f"Ошибка миграции: {e}")
        return False


def get_effective_balance_value(current_balance, balance_in_usd):
    usd_value = safe_float(balance_in_usd)
    if usd_value is not None and usd_value > 0:
        return usd_value
    current_value = safe_float(current_balance)
    return current_value if current_value is not None else 0.0


def get_effective_balance_history():
    rows = []
    if USE_DB:
        ensure_db_schema()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT date, {EFFECTIVE_BALANCE_SQL} FROM balances ORDER BY date ASC")
        raw_rows = cursor.fetchall()
        conn.close()
        for date_str, balance_value in raw_rows:
            if balance_value is None:
                continue
            try:
                rows.append((datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S'), float(balance_value)))
            except Exception:
                continue
        return rows

    worksheet_rows = list(worksheet.iter_rows(values_only=True))[1:]
    for row in worksheet_rows:
        if row[1] is None:
            continue
        try:
            rows.append((datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S'), float(row[1])))
        except Exception:
            continue
    return rows


def get_balance_at_or_before(rows, target_dt):
    chosen_value = None
    for row_dt, balance_value in rows:
        if row_dt <= target_dt:
            chosen_value = balance_value
        else:
            break
    return chosen_value


def get_closest_balance_value(rows, target_dt):
    closest_value = None
    closest_diff = None
    for row_dt, balance_value in rows:
        diff = abs((row_dt - target_dt).total_seconds())
        if closest_diff is None or diff < closest_diff:
            closest_diff = diff
            closest_value = balance_value
    return closest_value


def get_interval_slot_key(minutes, dt=None):
    interval = max(1, int(minutes))
    now_dt = dt or datetime.now()
    minute_stamp = int(now_dt.timestamp() // 60)
    slot_minute_stamp = minute_stamp - (minute_stamp % interval)
    return datetime.fromtimestamp(slot_minute_stamp * 60).strftime('%Y-%m-%d %H:%M:%S')


def claim_schedule_slot(slot_name, minutes, dt=None):
    slot_key = get_interval_slot_key(minutes, dt=dt)
    with SCHEDULER_LOCK:
        if SCHEDULER_STATE.get(slot_name) == slot_key:
            return False
        SCHEDULER_STATE[slot_name] = slot_key
    return True


def repair_balance_history(limit_rows=None):
    if not USE_DB:
        return {"deleted": 0, "updated": 0}

    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    if limit_rows is None:
        cursor.execute(
            "SELECT date, current_balance, balance_in_usd, balance_rub "
            "FROM balances ORDER BY date ASC"
        )
    else:
        cursor.execute(
            "SELECT date, current_balance, balance_in_usd, balance_rub "
            "FROM (SELECT date, current_balance, balance_in_usd, balance_rub "
            "FROM balances ORDER BY date DESC LIMIT ?) ORDER BY date ASC",
            (int(limit_rows),)
        )
    raw_rows = cursor.fetchall()

    parsed_rows = []
    for date_str, current_balance, balance_in_usd, balance_rub in raw_rows:
        try:
            row_dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        except Exception:
            continue
        effective_balance = get_effective_balance_value(current_balance, balance_in_usd)
        parsed_rows.append({
            "date": date_str,
            "dt": row_dt,
            "current_balance": safe_float(current_balance),
            "balance_in_usd": safe_float(balance_in_usd),
            "balance_rub": safe_float(balance_rub),
            "effective_balance": effective_balance
        })

    deleted_dates = []
    deduped_rows = []
    for row in parsed_rows:
        if deduped_rows:
            previous = deduped_rows[-1]
            same_minute = row["dt"].strftime('%Y-%m-%d %H:%M') == previous["dt"].strftime('%Y-%m-%d %H:%M')
            seconds_gap = (row["dt"] - previous["dt"]).total_seconds()
            previous_effective = previous["effective_balance"]
            current_effective = row["effective_balance"]
            balance_base = max((previous_effective + current_effective) / 2.0, 1.0)
            balance_diff_pct = abs(current_effective - previous_effective) / balance_base * 100.0
            if (same_minute and 0 <= seconds_gap <= BALANCE_REPAIR_DUPLICATE_SECONDS
                    and balance_diff_pct <= BALANCE_REPAIR_MAX_DUPLICATE_DIFF_PCT):
                deleted_dates.append(previous["date"])
                deduped_rows[-1] = row
                continue
        deduped_rows.append(row)

    updated_values = {}
    i = 1
    while i < len(deduped_rows) - 1:
        repaired_segment = False
        max_segment = min(BALANCE_REPAIR_MAX_SEGMENT_ROWS, len(deduped_rows) - i - 1)
        for segment_len in range(1, max_segment + 1):
            left_row = deduped_rows[i - 1]
            right_row = deduped_rows[i + segment_len]
            segment_rows = deduped_rows[i:i + segment_len]
            span_minutes = (right_row["dt"] - left_row["dt"]).total_seconds() / 60.0
            if span_minutes > BALANCE_REPAIR_MAX_SPAN_MINUTES:
                break

            boundary_avg = (left_row["effective_balance"] + right_row["effective_balance"]) / 2.0
            if boundary_avg <= 0:
                continue
            boundary_gap_pct = abs(right_row["effective_balance"] - left_row["effective_balance"]) / boundary_avg * 100.0
            min_segment_value = min(item["effective_balance"] for item in segment_rows)
            drop_usdt = boundary_avg - min_segment_value
            drop_pct = drop_usdt / boundary_avg * 100.0
            if boundary_gap_pct > BALANCE_REPAIR_MAX_BOUNDARY_GAP_PCT:
                continue
            if drop_pct < BALANCE_REPAIR_MIN_DROP_PCT or drop_usdt < BALANCE_REPAIR_MIN_DROP_USDT:
                continue
            if not all(item["effective_balance"] <= boundary_avg * BALANCE_REPAIR_SEGMENT_CEILING_RATIO
                       for item in segment_rows):
                continue

            for offset, segment_row in enumerate(segment_rows, start=1):
                corrected_value = left_row["effective_balance"] + (
                    (right_row["effective_balance"] - left_row["effective_balance"])
                    * (offset / (segment_len + 1))
                )
                updated_values[segment_row["date"]] = corrected_value
            i += segment_len
            repaired_segment = True
            break
        if not repaired_segment:
            i += 1

    for date_str in deleted_dates:
        cursor.execute("DELETE FROM balances WHERE date = ?", (date_str,))
        cursor.execute("DELETE FROM bot_snapshots WHERE snapshot_time = ?", (date_str,))

    for row in parsed_rows:
        corrected_value = updated_values.get(row["date"])
        if corrected_value is None:
            continue
        rub_balance = row["balance_rub"]
        effective_balance = row["effective_balance"]
        if rub_balance is not None and effective_balance not in (None, 0):
            rub_ratio = rub_balance / effective_balance
            corrected_rub = corrected_value * rub_ratio
        else:
            corrected_rub = rub_balance
        cursor.execute(
            "UPDATE balances SET current_balance = ?, balance_in_usd = ?, balance_rub = COALESCE(?, balance_rub) "
            "WHERE date = ?",
            (corrected_value, corrected_value, corrected_rub, row["date"])
        )

    conn.commit()
    conn.close()
    return {"deleted": len(deleted_dates), "updated": len(updated_values)}


# ------------------ ФУНКЦИИ ЗАПРОСА ДАННЫХ ------------------

BOT_LIST_URL = 'https://api2.bybit.com/s1/bot/tradingbot/v1/list-all-bots'
BOT_LIST_XAPI_URL = 'https://www.bybit.com/x-api/s1/bot/tradingbot/v1/list-all-bots'
BALANCE_URL = 'https://api2.bybit.com/v3/private/cht/asset-common/total-balance?quoteCoin=USDT&balanceType=1'
ASSET_SUMMARY_URL = "https://api2.bybit.com/bot-api-summary/v5/private/query-asset-summary"
MARK_PRICE_KLINE_URL = "https://api.bybit.com/v5/market/mark-price-kline"
MARKET_ALERT_WINDOW_SPECS = {
    30: {
        "calibration_interval": "30",
        "floor_pct": 0.9,
        "half_minutes": 15,
        "tail_minutes": 10,
        "half_ratio": 0.65,
        "tail_ratio": 0.35
    },
    60: {
        "calibration_interval": "60",
        "floor_pct": 1.4,
        "half_minutes": 30,
        "tail_minutes": 15,
        "half_ratio": 0.65,
        "tail_ratio": 0.35
    },
    120: {
        "calibration_interval": "120",
        "floor_pct": 2.1,
        "half_minutes": 60,
        "tail_minutes": 30,
        "half_ratio": 0.65,
        "tail_ratio": 0.35
    }
}
MARKET_ALERT_STATE = {
    "mute_until_ts": 0.0,
    "last_sent_minute_key": None,
    "calibration_cache": {}
}
MARKET_ALERT_MUTE_SECONDS = 1800
MARKET_ALERT_CALIBRATION_TTL_SECONDS = 6 * 3600
RISK_ALERT_STATE = {
    "mute_until_by_type": {}
}
RISK_ALERT_MUTE_SECONDS = 1800


def normalize_bybit_cookies(raw_value):
    if not raw_value:
        return {}

    if isinstance(raw_value, dict):
        normalized = {
            str(k).strip(): str(v).strip()
            for k, v in raw_value.items()
            if str(v).strip()
        }
        return normalized

    raw_text = str(raw_value).strip()
    if not raw_text:
        return {}

    if raw_text.startswith("{") and raw_text.endswith("}"):
        try:
            parsed_json = json.loads(raw_text)
            if isinstance(parsed_json, dict):
                normalized = {
                    str(k).strip(): str(v).strip()
                    for k, v in parsed_json.items()
                    if str(v).strip()
                }
                if normalized:
                    return normalized
        except Exception:
            pass

    cookie = SimpleCookie()
    try:
        cookie.load(raw_text)
    except Exception:
        cookie = SimpleCookie()

    parsed = {key: morsel.value for key, morsel in cookie.items() if morsel.value}
    if parsed:
        return parsed

    if "secure-token=" in raw_text:
        token_part = raw_text.split("secure-token=", 1)[1]
        token = token_part.split(";", 1)[0].strip()
        if token:
            return {"secure-token": token}

    return {"secure-token": raw_text}


def get_bybit_cookie_jar():
    jar = normalize_bybit_cookies(cookies)
    if not jar:
        return {}

    if not jar.get("secure-token"):
        for alt_key in ("secure_token", "secureToken", "securetoken"):
            alt_val = jar.get(alt_key)
            if alt_val:
                jar["secure-token"] = alt_val
                break

    return jar


def expire_mode_notify():
    global WAITING_FOR_RENEW
    WAITING_FOR_RENEW = True
    for admin_id in admins:
        try:
            bot.send_message(admin_id, "Срок действия данных истёк или возникла ошибка соединения. Обновите данные.")
        except Exception:
            pass


def retry_request(url, method='GET', headers=None, params=None, json_arg=None, cookies_arg=None, timeout=REQUEST_TIMEOUT,
                  notify_expire_on_fail=None, max_retries=None):
    if notify_expire_on_fail is None:
        notify_expire_on_fail = "bybit.com" in url
    if max_retries is None:
        max_retries = MAX_RETRIES

    attempts = 0
    while attempts < max_retries:
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers, params=params, cookies=cookies_arg, timeout=timeout)
            else:
                if json_arg is not None:
                    response = requests.post(url, headers=headers, json=json_arg, cookies=cookies_arg, timeout=timeout)
                else:
                    response = requests.post(url, headers=headers, data=params, cookies=cookies_arg, timeout=timeout)
            response.raise_for_status()
            return response
        except requests.RequestException as e:
            status_code = getattr(getattr(e, "response", None), "status_code", None)
            if status_code == 429 and not notify_expire_on_fail:
                logging.warning(f"Rate limit for {url}: {e}")
            else:
                logging.error(f"Ошибка запроса: {e}")
            attempts += 1
            sleep(2 ** attempts)
    if notify_expire_on_fail:
        expire_mode_notify()
    return None


def get_usdt_to_rub():
    now_ts = time.time()
    cached_value = RUB_CACHE["value"]
    updated_ts = RUB_CACHE["updated_ts"]

    if cached_value is not None and (now_ts - updated_ts) < RUB_CACHE["ttl_seconds"]:
        return cached_value

    response = retry_request(
        'https://api.coingecko.com/api/v3/simple/price?ids=tether&vs_currencies=rub',
        notify_expire_on_fail=False,
        max_retries=1
    )
    if response:
        data = response.json()
        try:
            value = float(data['tether']['rub'])
            RUB_CACHE["value"] = value
            RUB_CACHE["updated_ts"] = now_ts
            return value
        except (KeyError, ValueError, TypeError):
            pass

    if cached_value is not None and (now_ts - updated_ts) < RUB_CACHE["stale_ttl_seconds"]:
        return cached_value
    return None


def get_bybit_browser_headers(referer='https://www.bybit.com/'):
    return {
        'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                       'AppleWebKit/537.36 (KHTML, like Gecko) '
                       'Chrome/134.0.0.0 Safari/537.36'),
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.bybit.com',
        'Referer': referer,
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'Content-Type': 'application/json;charset=UTF-8'
    }


def fetch_bot_list_page(page_num=1, page_size=BOT_PAGE_SIZE, status=None):
    global WAITING_FOR_RENEW
    cookie_jar = get_bybit_cookie_jar()
    if not cookie_jar.get("secure-token"):
        expire_mode_notify()
        return [], 0

    payload = {
        "pageNum": int(page_num),
        "pageSize": int(page_size)
    }
    if status is not None:
        payload["status"] = int(status)

    response = retry_request(
        BOT_LIST_XAPI_URL,
        method='POST',
        headers=get_bybit_browser_headers(),
        json_arg=payload,
        cookies_arg=cookie_jar
    )
    if response:
        data = response.json()
        if data.get("ret_code") == 0:
            WAITING_FOR_RENEW = False
            result = data.get("result") or {}
            bots = result.get("bots") or []
            total = safe_int(result.get("total"))
            return bots, total if total is not None else len(bots)
        if data.get("ret_code") == 10007:
            expire_mode_notify()
    return [], 0


def fetch_all_bot_pages(status=None, page_size=BOT_PAGE_SIZE, max_pages=None):
    page_num = 1
    total = None
    all_bots = []

    while True:
        page_bots, page_total = fetch_bot_list_page(page_num=page_num, page_size=page_size, status=status)
        if total is None:
            total = page_total
        if not page_bots:
            break
        all_bots.extend(page_bots)
        if max_pages is not None and page_num >= max_pages:
            break
        if total is not None and len(all_bots) >= total:
            break
        if len(page_bots) < page_size:
            break
        page_num += 1

    return all_bots


def fetch_bot_list_data():
    return fetch_all_bot_pages()


def fetch_historical_bot_list_data():
    return fetch_all_bot_pages(status=BOT_HISTORY_STATUS)


def parse_total_balance_items(balance_items):
    account_balances = {}
    total_balance = 0.0
    has_total_balance = False

    for item in balance_items or []:
        item_balance = None
        for key in ('quoteBalance', 'originBalance'):
            parsed_value = safe_float((item or {}).get(key))
            if parsed_value is not None:
                item_balance = parsed_value
                break
        if item_balance is None:
            continue
        account_type = (item or {}).get('accountType') or 'ACCOUNT_TYPE_UNKNOWN'
        account_balances[account_type] = account_balances.get(account_type, 0.0) + item_balance
        total_balance += item_balance
        has_total_balance = True

    return account_balances, total_balance, has_total_balance


def fetch_balance_cookies(add_to_db=True):
    global WAITING_FOR_RENEW

    cookie_jar = get_bybit_cookie_jar()
    if not cookie_jar.get("secure-token"):
        expire_mode_notify()
        return "Не найден secure-token. Обновите cookies."

    response = retry_request(BALANCE_URL, cookies_arg=cookie_jar)
    if response:
        data = response.json()
        if data.get("ret_code") == 0:
            WAITING_FOR_RENEW = False
        if data.get("ret_code") == 10007:
            expire_mode_notify()
            return "Авторизация Bybit отклонена. Обновите cookies."

        balance_items = data.get('result', {}).get('totalBalanceItems', [])
        if not balance_items:
            expire_mode_notify()
            return "Срок действия cookies истёк. Бот в ожидании."

        account_balances, fallback_total_balance, has_fallback_total = parse_total_balance_items(balance_items)
        bot_wallet_balance = account_balances.get('ACCOUNT_TYPE_BOT', 0.0)
        funding_balance = account_balances.get('ACCOUNT_TYPE_FUND', 0.0)
        non_bot_balance = (
            max(0.0, fallback_total_balance - bot_wallet_balance)
            if has_fallback_total else 0.0
        )
        base_balance = fallback_total_balance if has_fallback_total else bot_wallet_balance
        active_bots = fetch_bot_list_data()
        active_bot_records = build_bot_archive_records(active_bots, is_active=True)

        bot_balance = bot_wallet_balance
        balance_in_usd = base_balance
        profit_in_usd = 0.0
        balance_in_btc = 0.0
        profit_in_btc = 0.0
        pnl_percentage = 0.0
        current_profit_in_usd = 0.0
        current_profit_in_btc = 0.0
        current_pnl_percentage = 0.0

        response2 = retry_request(ASSET_SUMMARY_URL, method='GET', cookies_arg=cookie_jar)
        if response2 is not None:
            data2 = response2.json()
            if data2.get("ret_code") == 0:
                asset_summary = data2["result"]["asset_summary"]
                try:
                    summary_balance = float(asset_summary.get("balance_in_usd", bot_wallet_balance))
                    if summary_balance > 0:
                        bot_balance = summary_balance
                except Exception:
                    bot_balance = bot_wallet_balance
                try:
                    profit_in_usd = float(asset_summary.get("profit_in_usd", 0))
                except Exception:
                    profit_in_usd = 0.0
                try:
                    balance_in_btc = float(asset_summary.get("balance_in_btc", 0))
                except Exception:
                    balance_in_btc = 0.0
                try:
                    profit_in_btc = float(asset_summary.get("profit_in_btc", 0))
                except Exception:
                    profit_in_btc = 0.0
                try:
                    pnl_percentage = float(asset_summary.get("pnl_percentage", 0))
                except Exception:
                    pnl_percentage = 0.0
                try:
                    current_profit_in_usd = float(asset_summary.get("current_profit_in_usd", 0))
                except Exception:
                    current_profit_in_usd = 0.0
                try:
                    current_profit_in_btc = float(asset_summary.get("current_profit_in_btc", 0))
                except Exception:
                    current_profit_in_btc = 0.0
                try:
                    current_pnl_percentage = float(asset_summary.get("current_pnl_percentage", 0))
                except Exception:
                    current_pnl_percentage = 0.0

        duplicate_record = None
        if has_fallback_total and active_bot_records:
            corrected_non_bot_balance, duplicate_record = correct_duplicate_non_bot_balance(
                non_bot_balance,
                active_bot_records
            )
            if corrected_non_bot_balance != non_bot_balance:
                logging.warning(
                    "Corrected duplicate funding jump for %s: %.2f -> %.2f USDT",
                    duplicate_record.get("symbol") if duplicate_record else "unknown",
                    non_bot_balance,
                    corrected_non_bot_balance
                )
                non_bot_balance = corrected_non_bot_balance
                funding_balance = min(funding_balance, non_bot_balance)

        if has_fallback_total:
            balance_in_usd = non_bot_balance + bot_balance
        else:
            balance_in_usd = bot_balance if bot_balance > 0 else base_balance

        current_balance = get_effective_balance_value(base_balance, balance_in_usd)
        origin_balance = bot_balance
        usdt_to_rub = get_usdt_to_rub()
        rub_balance = current_balance * usdt_to_rub if usdt_to_rub else 0.0
        now = datetime.now()
        history_rows = get_effective_balance_history()
        closest_balance_24h_ago = get_closest_balance_value(history_rows, now - timedelta(hours=24))

        if closest_balance_24h_ago is not None and closest_balance_24h_ago != 0:
            change_percent = ((current_balance - closest_balance_24h_ago) / closest_balance_24h_ago) * 100
        else:
            change_percent = 0

        now_str = now.strftime('%Y-%m-%d %H:%M:%S')
        if add_to_db:
            if USE_DB:
                ensure_db_schema()
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "REPLACE INTO balances (date, current_balance, balance_rub, change_percent, balance_in_usd, balance_in_btc, profit_in_usd, profit_in_btc, pnl_percentage, current_profit_in_usd, current_profit_in_btc, current_pnl_percentage, origin_balance, bot_balance, funding_balance, non_bot_balance, update_interval) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (now_str, current_balance, rub_balance, change_percent, balance_in_usd, balance_in_btc,
                     profit_in_usd, profit_in_btc, pnl_percentage, current_profit_in_usd,
                     current_profit_in_btc, current_pnl_percentage, origin_balance,
                     bot_balance, funding_balance, non_bot_balance,
                     config.get('db_update_interval', 30))
                )
                conn.commit()
                conn.close()
            else:
                worksheet.append([now_str, current_balance, rub_balance, change_percent])
                workbook.save(EXCEL_FILE)

            if USE_DB:
                try:
                    persist_bot_snapshots(now_str, active_bots)
                    persist_bot_archive_records(now_str, active_bots, is_active=True)
                    repair_balance_history(limit_rows=720)
                    repair_duplicate_bot_balance_spikes(limit_rows=720)
                except Exception as e:
                    logging.error(f"Ошибка сохранения истории ботов: {e}")

        sign = '🟢 +' if change_percent >= 0 else '🔴 '
        arrow = "📈" if change_percent >= 0 else "📉"
        change_str = f"{arrow} Изменение за 24ч: {sign}{change_percent:.2f}%"

        previous_message_balance = get_balance_at_or_before(
            history_rows,
            now - timedelta(minutes=max(1, int(balance_send_interval)))
        )
        if previous_message_balance is not None:
            diff_val = current_balance - previous_message_balance
            diff_sign = '🟢 +' if diff_val >= 0 else '🔴 '
            diff_str = f"Изменение к предыдущему сообщению: {diff_sign}{diff_val:.2f} USDT"
        else:
            diff_str = "Недостаточно данных для расчёта изменения к предыдущему сообщению."

        balance_info = (
            f"📅 Дата: {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"💰 Баланс аккаунта: {current_balance:.2f} USDT = {rub_balance:.2f} RUB\n"
            f"💵 Прибыль: {profit_in_usd:.2f} USDT\n"
            f"₿ Баланс BTC: {balance_in_btc:.8f}\n"
            f"🤖 Баланс ботов: {origin_balance:.2f} USDT\n"
            f"🏦 Вне ботов (Funding и др.): {non_bot_balance:.2f} USDT\n"
            f"{change_str}\n"
            f"{diff_str}"
        )
        return balance_info
    return "Ошибка соединения или данные недоступны"


def fetch_balance(add_to_db=True, bot_obj=None):
    return fetch_balance_cookies(add_to_db=add_to_db)


# ------------------ ФУНКЦИИ ДЛЯ ГРАФИКОВ ------------------

def percentile(values, q):
    if not values:
        return 0.0
    ordered = sorted(values)
    index = int((len(ordered) - 1) * q)
    index = max(0, min(index, len(ordered) - 1))
    return ordered[index]


def format_market_price(value):
    number = safe_float(value)
    if number is None:
        return "N/A"
    if number >= 100:
        return format_decimal(number, digits=2)
    if number >= 1:
        return format_decimal(number, digits=4)
    if number >= 0.01:
        return format_decimal(number, digits=5)
    return format_decimal(number, digits=6)


def fetch_public_mark_price_klines(symbol, interval, limit=200, start_ms=None, end_ms=None):
    params = {
        "category": "linear",
        "symbol": symbol,
        "interval": str(interval),
        "limit": limit
    }
    if start_ms is not None:
        params["start"] = int(start_ms)
    if end_ms is not None:
        params["end"] = int(end_ms)

    response = retry_request(
        MARK_PRICE_KLINE_URL,
        params=params,
        notify_expire_on_fail=False,
        max_retries=2
    )
    if not response:
        return []

    data = response.json()
    ret_code = data.get("retCode", data.get("ret_code"))
    if ret_code not in (0, None):
        return []

    rows = data.get("result", {}).get("list", []) or []
    klines = []
    for row in reversed(rows):
        try:
            klines.append({
                "ts": datetime.fromtimestamp(int(row[0]) / 1000),
                "open": float(row[1]),
                "high": float(row[2]),
                "low": float(row[3]),
                "close": float(row[4])
            })
        except Exception:
            continue
    return klines


def get_symbol_monthly_calibration(symbol):
    now_ts = time.time()
    cached = MARKET_ALERT_STATE["calibration_cache"].get(symbol)
    if cached and (now_ts - cached["updated_ts"]) < MARKET_ALERT_CALIBRATION_TTL_SECONDS:
        return cached["windows"]

    now_utc = datetime.now(timezone.utc)
    start_ms = int((now_utc - timedelta(days=30)).timestamp() * 1000)
    end_ms = int(now_utc.timestamp() * 1000)
    windows = {}

    for window_minutes, spec in MARKET_ALERT_WINDOW_SPECS.items():
        points = fetch_public_mark_price_klines(
            symbol,
            spec["calibration_interval"],
            limit=1000,
            start_ms=start_ms,
            end_ms=end_ms
        )
        closes = [item["close"] for item in points if item.get("close")]
        down_moves = []
        for previous_price, next_price in zip(closes, closes[1:]):
            if not previous_price:
                continue
            ret_pct = (next_price - previous_price) / previous_price * 100
            if ret_pct < 0:
                down_moves.append(-ret_pct)

        median_down = statistics.median(down_moves) if down_moves else 0.0
        q95_down = percentile(down_moves, 0.95) if down_moves else 0.0
        q98_down = percentile(down_moves, 0.98) if down_moves else 0.0
        threshold_pct = max(spec["floor_pct"], q95_down * 1.10, median_down * 2.40)
        windows[window_minutes] = {
            "threshold_pct": threshold_pct,
            "median_down_pct": median_down,
            "q95_down_pct": q95_down,
            "q98_down_pct": q98_down
        }

    MARKET_ALERT_STATE["calibration_cache"][symbol] = {
        "updated_ts": now_ts,
        "windows": windows
    }
    return windows


def compute_peak_drop(prices):
    if not prices:
        return 0.0, None
    peak_price = max(prices)
    current_price = prices[-1]
    if peak_price <= 0:
        return 0.0, peak_price
    return max(0.0, (peak_price - current_price) / peak_price * 100), peak_price


def analyze_symbol_market_state(symbol):
    calibration = get_symbol_monthly_calibration(symbol)
    minute_points = fetch_public_mark_price_klines(symbol, "1", limit=150)
    closes = [item["close"] for item in minute_points if item.get("close")]
    if len(closes) < 121:
        return None

    current_price = closes[-1]
    windows = {}
    for window_minutes, spec in MARKET_ALERT_WINDOW_SPECS.items():
        if len(closes) < (window_minutes + 1):
            continue

        window_prices = closes[-(window_minutes + 1):]
        half_prices = closes[-(spec["half_minutes"] + 1):]
        tail_prices = closes[-(spec["tail_minutes"] + 1):]

        window_drop_pct, peak_price = compute_peak_drop(window_prices)
        half_drop_pct, _ = compute_peak_drop(half_prices)
        tail_drop_pct, _ = compute_peak_drop(tail_prices)
        threshold_pct = calibration.get(window_minutes, {}).get("threshold_pct", spec["floor_pct"])

        triggered = (
            window_drop_pct >= threshold_pct
            and half_drop_pct >= max(threshold_pct * spec["half_ratio"], window_drop_pct * 0.55)
            and tail_drop_pct >= max(threshold_pct * spec["tail_ratio"], window_drop_pct * 0.30)
        )

        windows[window_minutes] = {
            "triggered": triggered,
            "threshold_pct": threshold_pct,
            "window_drop_pct": window_drop_pct,
            "half_drop_pct": half_drop_pct,
            "tail_drop_pct": tail_drop_pct,
            "peak_price": peak_price,
            "current_price": current_price
        }

    return {
        "symbol": symbol,
        "current_price": current_price,
        "windows": windows,
        "calibration": calibration
    }


def get_drop_sensitive_bot_snapshots():
    bots_data = fetch_bot_list_data()[:6]
    snapshots = []
    for index, bot_data in enumerate(bots_data):
        snapshot = build_bot_snapshot(bot_data, index)
        symbol = snapshot.get("symbol")
        if not symbol:
            continue
        badge = str(snapshot.get("badge") or "").lower()
        if "short" in badge:
            continue
        snapshots.append(snapshot)
    return snapshots


def evaluate_market_drop_signal():
    bot_snapshots = get_drop_sensitive_bot_snapshots()
    unique_symbols = sorted({snapshot["symbol"] for snapshot in bot_snapshots if snapshot.get("symbol")})
    if len(unique_symbols) < 2:
        return None

    symbol_states = {}
    common_windows = None
    for symbol in unique_symbols:
        state = analyze_symbol_market_state(symbol)
        if not state:
            return None
        symbol_states[symbol] = state
        triggered_windows = {window for window, meta in state["windows"].items() if meta.get("triggered")}
        if not triggered_windows:
            return None
        common_windows = triggered_windows if common_windows is None else common_windows & triggered_windows
        if not common_windows:
            return None

    selected_window = min(common_windows)
    total_invested = sum(snapshot.get("investment_usdt") or 0.0 for snapshot in bot_snapshots)
    total_equity = sum(
        snapshot.get("equity_usdt")
        if snapshot.get("equity_usdt") is not None
        else (snapshot.get("investment_usdt") or 0.0)
        for snapshot in bot_snapshots
    )
    total_pnl = sum(snapshot.get("pnl_usdt") or 0.0 for snapshot in bot_snapshots)
    severity = min(
        symbol_states[symbol]["windows"][selected_window]["window_drop_pct"] /
        max(symbol_states[symbol]["windows"][selected_window]["threshold_pct"], 0.0001)
        for symbol in unique_symbols
    )

    return {
        "window_minutes": selected_window,
        "severity": severity,
        "symbols": unique_symbols,
        "symbol_states": symbol_states,
        "bot_snapshots": bot_snapshots,
        "total_invested": total_invested,
        "total_equity": total_equity,
        "total_pnl": total_pnl
    }


def build_market_alert_markup():
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Отключить сигнализацию на 30 мин", callback_data="alert_mute_30m"))
    return markup


def build_market_alert_message(alert_state):
    window_minutes = alert_state["window_minutes"]
    lines = [
        "Резкое синхронное падение рынка",
        f"Окно: {window_minutes} мин | сила сигнала: x{alert_state['severity']:.2f}",
        (
            f"Портфель ботов: вложено {format_usdt(alert_state['total_invested'])} | "
            f"текущая стоимость {format_usdt(alert_state['total_equity'])} | "
            f"P&L {format_usdt(alert_state['total_pnl'])}"
        ),
        "",
        "Монеты:"
    ]

    for symbol in alert_state["symbols"]:
        symbol_meta = alert_state["symbol_states"][symbol]["windows"][window_minutes]
        lines.append(
            f"{symbol}: -{symbol_meta['window_drop_pct']:.2f}% "
            f"от локального пика {format_market_price(symbol_meta['peak_price'])} "
            f"до {format_market_price(symbol_meta['current_price'])} "
            f"(порог {symbol_meta['threshold_pct']:.2f}%)"
        )

    lines.append("")
    lines.append("Боты:")
    for snapshot in alert_state["bot_snapshots"][:6]:
        current_value = (
            snapshot["equity_usdt"]
            if snapshot.get("equity_usdt") is not None
            else (snapshot.get("investment_usdt") or 0.0)
        )
        lines.append(
            f"{snapshot['title']} {snapshot['badge']}: "
            f"стоимость {format_usdt(current_value)} | "
            f"вложено {format_usdt(snapshot.get('investment_usdt'))} | "
            f"P&L {format_usdt(snapshot.get('pnl_usdt'))}"
        )

    lines.append("")
    lines.append("Если вы контролируете ситуацию, отключите сигнализацию на 30 минут.")
    return "\n".join(lines)


def send_market_alert(alert_state):
    if not admins:
        return

    minute_key = datetime.now().strftime("%Y-%m-%d %H:%M")
    if MARKET_ALERT_STATE["last_sent_minute_key"] == minute_key:
        return

    message_text = build_market_alert_message(alert_state)
    markup = build_market_alert_markup()
    for admin_id in admins:
        try:
            bot.send_message(admin_id, message_text, reply_markup=markup)
        except Exception as e:
            logging.error(f"Ошибка отправки market alert админу {admin_id}: {e}")

    MARKET_ALERT_STATE["last_sent_minute_key"] = minute_key


def check_market_alerts():
    if not get_notification_settings().get("market_drop", True):
        return
    if WAITING_FOR_RENEW or not get_bybit_cookie_jar().get("secure-token"):
        return

    if time.time() < MARKET_ALERT_STATE["mute_until_ts"]:
        return

    alert_state = evaluate_market_drop_signal()
    if not alert_state:
        MARKET_ALERT_STATE["last_sent_minute_key"] = None
        return

    send_market_alert(alert_state)


def market_alert_loop(run_token):
    while not stop_threads and run_token == thread_run_token:
        try:
            check_market_alerts()
        except Exception as e:
            logging.error(f"Ошибка цикла market alert: {e}")
        if not wait_until_next_interval(1, run_token=run_token):
            break


# ------------------ ФУНКЦИИ ДЛЯ ГРАФИКОВ ------------------
def format_duration(sec_str):
    seconds = int(sec_str)
    days = seconds // 86400
    seconds %= 86400
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    return f"{days}D {hours}h {minutes}m"


def get_all_dates():
    if USE_DB:
        ensure_db_schema()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT date FROM balances ORDER BY date ASC")
        rows = cursor.fetchall()
        conn.close()
        dates = []
        for r in rows:
            try:
                d = datetime.strptime(r[0], '%Y-%m-%d %H:%M:%S').date()
                dates.append(d)
            except Exception:
                continue
        return sorted(list(set(dates)))
    else:
        rows = list(worksheet.iter_rows(values_only=True))[1:]
        dates = sorted(list(set(datetime.strptime(r[0], '%Y-%m-%d %H:%M:%S').date() for r in rows)))
        return dates


def get_months_from_dates(dates):
    ym_set = set((d.year, d.month) for d in dates)
    return sorted(ym_set)


def dates_in_month(dates, year, month):
    return [d for d in dates if d.year == year and d.month == month]


def month_name(year, month):
    return datetime(year, month, 1).strftime('%B %Y')


def generate_calendar_markup(selected_year, selected_month):
    dates = get_all_dates()
    if not dates:
        return None, (selected_year, selected_month)
    months = get_months_from_dates(dates)
    if (selected_year, selected_month) not in months:
        selected_year, selected_month = months[-1]
    current_month_dates = [d for d in dates if d.year == selected_year and d.month == selected_month]
    markup = types.InlineKeyboardMarkup(row_width=7)
    day_buttons = []
    for d in current_month_dates:
        day_str = f"{d.day:02d}"
        cb_data = f"graph_day_{d.strftime('%d_%m_%Y')}"
        day_buttons.append(types.InlineKeyboardButton(day_str, callback_data=cb_data))
    if day_buttons:
        markup.add(*day_buttons)
    idx = months.index((selected_year, selected_month))
    prev_month_cb = None
    next_month_cb = None
    if idx > 0:
        py, pm = months[idx - 1]
        prev_month_cb = f"graph_month_{py}_{pm:02d}"
    if idx < len(months) - 1:
        ny, nm = months[idx + 1]
        next_month_cb = f"graph_month_{ny}_{nm:02d}"
    nav_buttons = []
    if prev_month_cb:
        nav_buttons.append(
            types.InlineKeyboardButton("<", callback_data=f"graph_monthnav_prev_{selected_year}_{selected_month:02d}"))
    nav_buttons.append(types.InlineKeyboardButton(month_name(selected_year, selected_month),
                                                  callback_data=f"graph_month_{selected_year}_{selected_month:02d}"))
    if next_month_cb:
        nav_buttons.append(
            types.InlineKeyboardButton(">", callback_data=f"graph_monthnav_next_{selected_year}_{selected_month:02d}"))
    markup.add(*nav_buttons)
    return markup, (selected_year, selected_month)


def get_default_month():
    dates = get_all_dates()
    if not dates:
        return None
    months = get_months_from_dates(dates)
    return months[-1] if months else None


def safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_int(value):
    try:
        if value in (None, ""):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_epoch_timestamp(value):
    ts_value = safe_int(value)
    if ts_value is None:
        return None
    if ts_value > 10 ** 12:
        return ts_value // 1000
    return ts_value


def format_short_datetime(ts_value, fallback="N/A"):
    normalized = normalize_epoch_timestamp(ts_value)
    if normalized is None:
        return fallback
    return datetime.fromtimestamp(normalized).strftime('%d.%m %H:%M')


def format_duration_from_timestamps(start_ts, end_ts=None, fallback="N/A"):
    start_value = normalize_epoch_timestamp(start_ts)
    end_value = normalize_epoch_timestamp(end_ts)
    if start_value is None:
        return fallback
    if end_value is None:
        end_value = int(time.time())
    if end_value < start_value:
        return fallback
    return format_duration(str(end_value - start_value))


def parse_usdt_amounts(text):
    if not text:
        return []
    return [
        float(match)
        for match in re.findall(r'([+-]?\d+(?:\.\d+)?)\s*USDT', str(text), flags=re.IGNORECASE)
    ]


def extract_settlement_assets_usdt(text):
    amounts = parse_usdt_amounts(text)
    if not amounts:
        return None
    return sum(amounts)


def get_bot_detail_payload(bot_data):
    bot_type = (bot_data or {}).get("type", "UNKNOWN")
    future_grid = (bot_data or {}).get("future_grid") or {}
    futures_mart = (bot_data or {}).get("fmart") or {}
    spot_grid = (((bot_data or {}).get("grid") or {}).get("info")) or {}
    spot_profit = (((bot_data or {}).get("grid") or {}).get("profit")) or {}
    combo = (bot_data or {}).get("fcombo") or {}
    return bot_type, future_grid, futures_mart, spot_grid, spot_profit, combo


def infer_bot_active_flag(status_value, display_status_value, explicit_is_active=None):
    if explicit_is_active is not None:
        return int(bool(explicit_is_active))
    status_blob = " ".join(
        str(item or "") for item in (status_value, display_status_value)
    ).upper()
    if not status_blob:
        return 0
    if "COMPLETED" in status_blob or "CLOSED" in status_blob or "STOP" in status_blob:
        return 0
    if "RUNNING" in status_blob or "ACTIVE" in status_blob:
        return 1
    return 0


def get_bot_identity_key(bot_id=None, symbol=None, bot_type=None):
    if bot_id:
        return str(bot_id)
    return f"{symbol or 'UNKNOWN'}|{bot_type or 'UNKNOWN'}"


def calculate_profit_percent(profit_value, investment_value):
    profit_number = safe_float(profit_value)
    investment_number = safe_float(investment_value)
    if profit_number is None or investment_number in (None, 0):
        return None
    return profit_number / investment_number * 100.0


def build_bot_archive_record(bot_data, is_active=None):
    snapshot = build_bot_snapshot(bot_data, 0)
    bot_type, future_grid, futures_mart, spot_grid, spot_profit, combo = get_bot_detail_payload(bot_data)

    bot_id = None
    status_value = None
    display_status_value = None
    close_code = None
    close_reason = None
    settlement_assets_text = None
    settlement_assets_usdt = None
    leverage = None
    mode = None
    created_ts = None
    ended_ts = None

    if bot_type == "GRID_FUTURES" and future_grid:
        bot_id = future_grid.get("grid_id")
        status_value = future_grid.get("status")
        display_status_value = status_value
        close_detail = future_grid.get("close_detail") or {}
        close_code = close_detail.get("bot_close_code")
        close_reason = close_detail.get("close_reason")
        settlement_assets_text = close_detail.get("settlement_asset")
        leverage = future_grid.get("leverage")
        mode = future_grid.get("grid_mode")
        created_ts = normalize_epoch_timestamp(future_grid.get("create_time"))
        ended_ts = normalize_epoch_timestamp(future_grid.get("end_time"))
    elif bot_type == "MART_FUTURES" and futures_mart:
        bot_id = futures_mart.get("bot_id")
        status_value = futures_mart.get("bot_display_status")
        display_status_value = status_value
        close_code = futures_mart.get("close_code")
        close_reason = futures_mart.get("stop_type")
        settlement_assets_text = futures_mart.get("settlement_assets")
        leverage = futures_mart.get("leverage")
        mode = futures_mart.get("fmart_mode")
        created_ts = normalize_epoch_timestamp(futures_mart.get("create_time"))
        ended_ts = normalize_epoch_timestamp(futures_mart.get("end_time"))
    elif bot_type == "GRID_SPOT" and spot_grid:
        bot_id = spot_grid.get("grid_id")
        status_value = spot_grid.get("status")
        display_status_value = status_value
        close_code = spot_grid.get("bot_close_code")
        close_reason = spot_grid.get("close_reason")
        settlement_assets_text = spot_profit.get("settlement_assets")
        leverage = None
        mode = spot_grid.get("grid_mode")
        created_ts = normalize_epoch_timestamp(spot_grid.get("create_time"))
        ended_ts = normalize_epoch_timestamp(spot_grid.get("modify_time"))
    elif bot_type == "COMBO_FUTURES" and combo:
        bot_id = combo.get("bot_id")
        status_value = combo.get("bot_display_status")
        display_status_value = status_value
        close_code = combo.get("close_code")
        close_reason = combo.get("stop_type")
        settlement_assets_text = combo.get("settlement_assets")
        leverage = combo.get("leverage")
        mode = combo.get("bot_mode")
        created_ts = normalize_epoch_timestamp(combo.get("create_time"))
        ended_ts = normalize_epoch_timestamp(combo.get("end_time"))

    settlement_assets_usdt = extract_settlement_assets_usdt(settlement_assets_text)
    active_flag = infer_bot_active_flag(status_value, display_status_value, explicit_is_active=is_active)
    pnl_usdt = snapshot.get("pnl_usdt")
    investment_usdt = snapshot.get("investment_usdt")
    final_profit_usdt = pnl_usdt if not active_flag else None
    if final_profit_usdt is None and not active_flag and settlement_assets_usdt is not None and investment_usdt is not None:
        final_profit_usdt = settlement_assets_usdt - investment_usdt
    normalized_profit_usdt = pnl_usdt if active_flag else (final_profit_usdt if final_profit_usdt is not None else pnl_usdt)
    normalized_percent = calculate_profit_percent(normalized_profit_usdt, investment_usdt)
    if normalized_percent is None:
        normalized_percent = snapshot.get("pnl_percent_value")

    return {
        "bot_id": str(bot_id) if bot_id not in (None, "") else None,
        "identity_key": get_bot_identity_key(bot_id, snapshot.get("symbol"), bot_type),
        "symbol": snapshot.get("symbol"),
        "bot_type": bot_type,
        "title": snapshot.get("title"),
        "badge": snapshot.get("badge"),
        "status": status_value,
        "display_status": display_status_value,
        "close_code": close_code,
        "close_reason": close_reason,
        "investment_usdt": investment_usdt,
        "pnl_usdt": pnl_usdt,
        "equity_usdt": snapshot.get("equity_usdt"),
        "pnl_percent": normalized_percent,
        "final_profit_usdt": final_profit_usdt,
        "settlement_assets_text": settlement_assets_text,
        "settlement_assets_usdt": settlement_assets_usdt,
        "leverage": leverage,
        "mode": mode,
        "created_ts": created_ts,
        "ended_ts": ended_ts,
        "is_active": active_flag,
        "raw_json": json.dumps(bot_data, ensure_ascii=False)
    }


def build_bot_archive_records(bots_data, is_active=None):
    records = []
    for bot_data in bots_data or []:
        if not isinstance(bot_data, dict):
            continue
        records.append(build_bot_archive_record(bot_data, is_active=is_active))
    return records


def persist_bot_archive_records(snapshot_time, bots_data, is_active=None):
    if not USE_DB or not bots_data:
        return 0
    ensure_db_schema()
    records = build_bot_archive_records(bots_data, is_active=is_active)
    conn = get_db_connection()
    cursor = conn.cursor()
    saved = 0
    for record in records:
        bot_id = record.get("bot_id")
        if not bot_id:
            continue
        cursor.execute(
            """
            INSERT INTO bot_archive (
                bot_id, symbol, bot_type, title, badge, status, display_status, close_code, close_reason,
                investment_usdt, pnl_usdt, equity_usdt, pnl_percent, final_profit_usdt,
                settlement_assets_text, settlement_assets_usdt, leverage, mode,
                created_ts, ended_ts, first_seen_at, last_seen_at, last_snapshot_time, is_active, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(bot_id) DO UPDATE SET
                symbol = excluded.symbol,
                bot_type = excluded.bot_type,
                title = excluded.title,
                badge = excluded.badge,
                status = excluded.status,
                display_status = excluded.display_status,
                close_code = COALESCE(excluded.close_code, bot_archive.close_code),
                close_reason = COALESCE(excluded.close_reason, bot_archive.close_reason),
                investment_usdt = excluded.investment_usdt,
                pnl_usdt = excluded.pnl_usdt,
                equity_usdt = excluded.equity_usdt,
                pnl_percent = excluded.pnl_percent,
                final_profit_usdt = COALESCE(excluded.final_profit_usdt, bot_archive.final_profit_usdt),
                settlement_assets_text = COALESCE(excluded.settlement_assets_text, bot_archive.settlement_assets_text),
                settlement_assets_usdt = COALESCE(excluded.settlement_assets_usdt, bot_archive.settlement_assets_usdt),
                leverage = COALESCE(excluded.leverage, bot_archive.leverage),
                mode = COALESCE(excluded.mode, bot_archive.mode),
                created_ts = COALESCE(excluded.created_ts, bot_archive.created_ts),
                ended_ts = COALESCE(excluded.ended_ts, bot_archive.ended_ts),
                first_seen_at = COALESCE(bot_archive.first_seen_at, excluded.first_seen_at),
                last_seen_at = excluded.last_seen_at,
                last_snapshot_time = excluded.last_snapshot_time,
                is_active = excluded.is_active,
                raw_json = excluded.raw_json
            """,
            (
                bot_id,
                record.get("symbol"),
                record.get("bot_type"),
                record.get("title"),
                record.get("badge"),
                record.get("status"),
                record.get("display_status"),
                record.get("close_code"),
                record.get("close_reason"),
                record.get("investment_usdt"),
                record.get("pnl_usdt"),
                record.get("equity_usdt"),
                record.get("pnl_percent"),
                record.get("final_profit_usdt"),
                record.get("settlement_assets_text"),
                record.get("settlement_assets_usdt"),
                record.get("leverage"),
                record.get("mode"),
                record.get("created_ts"),
                record.get("ended_ts"),
                snapshot_time,
                snapshot_time,
                snapshot_time,
                record.get("is_active"),
                record.get("raw_json")
            )
        )
        saved += 1
    conn.commit()
    conn.close()
    return saved


def sync_bot_archive(force=False, include_active=False, include_history=True):
    if not USE_DB:
        return 0
    if not force and not claim_schedule_slot("bot_archive_slot", BOT_ARCHIVE_SYNC_INTERVAL_MINUTES):
        return 0

    cookie_jar = get_bybit_cookie_jar()
    if not cookie_jar.get("secure-token"):
        return 0

    snapshot_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    saved = 0
    try:
        if include_active:
            active_bots = fetch_bot_list_data()
            saved += persist_bot_archive_records(snapshot_time, active_bots, is_active=True)
        if include_history:
            history_bots = fetch_historical_bot_list_data()
            saved += persist_bot_archive_records(snapshot_time, history_bots, is_active=False)
    except Exception as e:
        logging.error(f"Ошибка синхронизации архива ботов: {e}")
    return saved


def repair_bot_archive_metrics():
    if not USE_DB:
        return 0

    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT bot_id, investment_usdt, pnl_usdt, final_profit_usdt, settlement_assets_usdt, is_active, pnl_percent
        FROM bot_archive
        """
    )
    rows = cursor.fetchall()
    updated = 0
    for bot_id, investment_usdt, pnl_usdt, final_profit_usdt, settlement_assets_usdt, is_active, pnl_percent in rows:
        investment_number = safe_float(investment_usdt)
        pnl_number = safe_float(pnl_usdt)
        final_profit_number = safe_float(final_profit_usdt)
        settlement_number = safe_float(settlement_assets_usdt)
        active_flag = bool(safe_int(is_active))

        expected_final_profit = None
        if not active_flag:
            expected_final_profit = pnl_number
            if expected_final_profit is None and investment_number is not None and settlement_number is not None:
                expected_final_profit = settlement_number - investment_number

        expected_profit_for_percent = pnl_number if active_flag else (
            expected_final_profit if expected_final_profit is not None else pnl_number
        )
        expected_percent = calculate_profit_percent(expected_profit_for_percent, investment_number)

        current_percent = safe_float(pnl_percent)
        final_changed = (not active_flag and expected_final_profit != final_profit_number)
        percent_changed = (
            (expected_percent is None and current_percent is not None)
            or (expected_percent is not None and current_percent is None)
            or (
                expected_percent is not None and current_percent is not None
                and abs(expected_percent - current_percent) > 1e-9
            )
        )

        if not final_changed and not percent_changed:
            continue

        cursor.execute(
            """
            UPDATE bot_archive
            SET final_profit_usdt = ?, pnl_percent = ?
            WHERE bot_id = ?
            """,
            (
                expected_final_profit if not active_flag else None,
                expected_percent,
                bot_id
            )
        )
        updated += 1

    conn.commit()
    conn.close()
    return updated


def classify_bot_close_notification_type(record):
    close_code = str((record or {}).get("close_code") or "").upper()
    close_reason = str((record or {}).get("close_reason") or "").upper()
    status_blob = " ".join([close_code, close_reason])
    if "LIQ" in status_blob:
        return "bot_liquidation"
    if "AUTO_SL" in status_blob or "BY_SL" in status_blob or "STOP_LOSS" in status_blob:
        return "bot_stop_loss"
    if "TRAILING" in status_blob:
        return "bot_trailing_stop"
    if "MANUALLY" in status_blob or "BY_USER" in status_blob or "USER" in status_blob:
        return "bot_manual_close"
    return None


def mark_bot_close_notified(bot_id, notify_type):
    if not USE_DB or not bot_id:
        return
    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE bot_archive
        SET close_notified_at = ?, close_notify_type = ?
        WHERE bot_id = ?
        """,
        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), notify_type, str(bot_id))
    )
    conn.commit()
    conn.close()


def bootstrap_bot_close_notifications():
    if not USE_DB or config.get("bot_close_notify_bootstrapped"):
        return 0
    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute(
        """
        UPDATE bot_archive
        SET close_notified_at = COALESCE(close_notified_at, ?),
            close_notify_type = COALESCE(close_notify_type, 'bootstrap')
        WHERE COALESCE(is_active, 0) = 0
          AND close_notified_at IS NULL
        """,
        (now_str,)
    )
    updated = cursor.rowcount
    conn.commit()
    conn.close()
    config["bot_close_notify_bootstrapped"] = True
    save_config(config)
    return updated


def build_bot_close_notification_message(record, notify_type):
    label = NOTIFICATION_LABELS.get(notify_type, "Закрытие бота")
    symbol = record.get("symbol") or record.get("title") or "UNKNOWN"
    bot_type = get_bot_type_name(record.get("bot_type"))
    investment_value = safe_float(record.get("investment_usdt"))
    profit_value = get_top_bot_profit_value(record)
    percent_value = get_top_bot_percent_value(record)
    close_time = format_short_datetime(record.get("ended_ts"), fallback=record.get("last_snapshot_time") or "N/A")
    badge = record.get("badge") or record.get("mode") or "N/A"
    close_code = record.get("close_code") or "N/A"
    close_reason = record.get("close_reason") or "N/A"
    return (
        f"Уведомление: {label}\n"
        f"Бот: {symbol}\n"
        f"Тип: {bot_type}\n"
        f"Режим: {badge}\n"
        f"Вложение: {format_usdt(investment_value, fallback='N/A')}\n"
        f"Итог: {format_usdt(profit_value, fallback='N/A')} / {format_percent(percent_value, scale=1, fallback='N/A')}\n"
        f"Закрыт: {close_time}\n"
        f"Причина: {close_code}\n"
        f"Деталь: {close_reason}\n"
        f"ID: {record.get('bot_id') or 'N/A'}"
    )


def dispatch_bot_close_notifications(limit=20):
    if not USE_DB or not admins:
        return 0

    settings = get_notification_settings()
    ensure_db_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT *
        FROM bot_archive
        WHERE COALESCE(is_active, 0) = 0
          AND close_notified_at IS NULL
        ORDER BY COALESCE(ended_ts, created_ts, 0) ASC
        LIMIT ?
        """,
        (int(limit),)
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    notified = 0
    for record in rows:
        notify_type = classify_bot_close_notification_type(record)
        if notify_type is None:
            mark_bot_close_notified(record.get("bot_id"), "ignored")
            continue
        if not settings.get(notify_type, False):
            mark_bot_close_notified(record.get("bot_id"), f"disabled:{notify_type}")
            continue

        message_text = build_bot_close_notification_message(record, notify_type)
        sent = False
        for admin_id in admins:
            try:
                bot.send_message(admin_id, message_text)
                sent = True
            except Exception as e:
                logging.error(f"Ошибка отправки close alert админу {admin_id}: {e}")
        if sent:
            mark_bot_close_notified(record.get("bot_id"), notify_type)
            notified += 1
    return notified


def derive_direction(mode_value=None, badge_value=None):
    blob = " ".join(str(item or "") for item in (mode_value, badge_value)).upper()
    if "SHORT" in blob:
        return "short"
    if "LONG" in blob:
        return "long"
    if "NEUTRAL" in blob:
        return "neutral"
    return "unknown"


def parse_leverage_value(value):
    number = safe_float(value)
    if number is not None:
        return number
    match = re.search(r'(\d+(?:\.\d+)?)', str(value or ""))
    return safe_float(match.group(1)) if match else None


def get_alert_event(alert_key):
    if not USE_DB:
        return None
    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT alert_key, created_at FROM alert_events WHERE alert_key = ?",
        (str(alert_key),)
    )
    row = cursor.fetchone()
    conn.close()
    return row


def record_alert_event(alert_key, alert_type, bot_id=None, symbol=None, payload=None):
    if not USE_DB:
        return
    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT OR REPLACE INTO alert_events (alert_key, alert_type, bot_id, symbol, created_at, payload_json)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            str(alert_key),
            alert_type,
            str(bot_id) if bot_id not in (None, "") else None,
            symbol,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            json.dumps(payload or {}, ensure_ascii=False)
        )
    )
    conn.commit()
    conn.close()


def get_bot_initial_snapshot_metrics(bot_id):
    if not USE_DB or not bot_id:
        return None
    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT snapshot_time, investment_usdt, pnl_usdt
        FROM bot_snapshots
        WHERE bot_id = ?
        ORDER BY snapshot_time ASC
        LIMIT 1
        """,
        (str(bot_id),)
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    return {
        "snapshot_time": row[0],
        "investment_usdt": safe_float(row[1]),
        "pnl_usdt": safe_float(row[2])
    }


def find_recent_symbol_loss(symbol, cooldown_hours):
    if not USE_DB or not symbol:
        return None
    ensure_db_schema()
    cooldown_seconds = max(0.0, safe_float(cooldown_hours) or 0.0) * 3600.0
    ended_after = int(time.time() - cooldown_seconds)
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT *
        FROM bot_archive
        WHERE COALESCE(is_active, 0) = 0
          AND symbol = ?
          AND COALESCE(ended_ts, created_ts, 0) >= ?
          AND COALESCE(final_profit_usdt, pnl_usdt, 0) < 0
        ORDER BY COALESCE(ended_ts, created_ts, 0) DESC
        LIMIT 1
        """,
        (symbol, int(ended_after))
    )
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def build_active_risk_record(bot_data, index):
    snapshot = build_bot_snapshot(bot_data, index)
    archive_record = build_bot_archive_record(bot_data, is_active=True)
    bot_type, future_grid, futures_mart, spot_grid, _, combo = get_bot_detail_payload(bot_data)
    mark_price = None
    liq_price = None
    position_count = None
    if bot_type == "GRID_FUTURES" and future_grid:
        mark_price = safe_float(future_grid.get("mark_price"))
        liq_price = safe_float(future_grid.get("liq_price"))
        position_count = safe_int(future_grid.get("arbitrage_num"))
    elif bot_type == "MART_FUTURES" and futures_mart:
        mark_price = safe_float(futures_mart.get("current_price"))
        liq_price = safe_float(futures_mart.get("liq_price"))
        position_count = safe_int(futures_mart.get("added_pos_num"))
    elif bot_type == "COMBO_FUTURES" and combo:
        mark_price = safe_float(combo.get("mark_price"))
        liq_price = safe_float(combo.get("liq_price"))
        position_count = safe_int(combo.get("position_num"))

    leverage_value = parse_leverage_value(archive_record.get("leverage") or snapshot.get("badge"))
    pnl_percent_value = calculate_profit_percent(snapshot.get("pnl_usdt"), snapshot.get("investment_usdt"))
    if pnl_percent_value is None:
        pnl_percent_value = snapshot.get("pnl_percent_value")

    return {
        "bot_id": archive_record.get("bot_id"),
        "symbol": archive_record.get("symbol") or snapshot.get("symbol"),
        "bot_type": bot_type,
        "title": snapshot.get("title"),
        "badge": snapshot.get("badge"),
        "status_text": snapshot.get("status_text"),
        "started_at_text": snapshot.get("started_at_text"),
        "grid_cells_mode": snapshot.get("grid_cells_mode"),
        "mode": archive_record.get("mode"),
        "direction": derive_direction(archive_record.get("mode"), snapshot.get("badge")),
        "investment_usdt": snapshot.get("investment_usdt"),
        "pnl_usdt": snapshot.get("pnl_usdt"),
        "pnl_percent": pnl_percent_value,
        "realized_pnl_usdt": snapshot.get("realized_pnl_usdt"),
        "unrealized_pnl_usdt": snapshot.get("unrealized_pnl_usdt"),
        "leverage": leverage_value,
        "mark_price": mark_price,
        "liq_price": liq_price,
        "position_count": position_count,
        "current_price": snapshot.get("current_price"),
        "entry_price": snapshot.get("entry_price"),
        "tp_sl_text": snapshot.get("tp_sl_text"),
        "trailing_text": snapshot.get("trailing_text"),
        "arbitrage_num": snapshot.get("arbitrage_num")
    }


def build_risk_alert_message(alert_type, record, extra_text):
    title_map = {
        "bot_risk_limit": "Нарушение риск-лимита",
        "bot_pnl_drawdown": "Бот ушёл в глубокую просадку",
        "bot_liq_distance": "Бот близко к ликвидации",
        "bot_repeat_loss": "Повторный вход после свежего лося"
    }
    lines = [
        f"Риск-контроль: {title_map.get(alert_type, alert_type)}",
        f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Бот: {record.get('symbol') or record.get('title')}",
        f"Тип: {get_bot_type_name(record.get('bot_type'))}",
        f"Режим: {record.get('badge') or record.get('mode') or 'N/A'}",
        f"Статус: {record.get('status_text') or 'N/A'}",
        f"Время запуска: {record.get('started_at_text') or 'N/A'}",
        f"Общая сумма инвестиций: {format_usdt(record.get('investment_usdt'), fallback='N/A')}",
        f"Общий P&L: {format_usdt(record.get('pnl_usdt'), fallback='N/A')}",
        f"Доходность: {format_percent(record.get('pnl_percent'), scale=1, fallback='N/A')}"
    ]
    if record.get("realized_pnl_usdt") is not None:
        lines.append(f"Полученная прибыль: {format_usdt(record.get('realized_pnl_usdt'), fallback='N/A')}")
    if record.get("unrealized_pnl_usdt") is not None:
        lines.append(f"Нереализованный P&L: {format_usdt(record.get('unrealized_pnl_usdt'), fallback='N/A')}")
    if record.get("grid_cells_mode") not in (None, "", "N/A"):
        lines.append(f"Сетки (режим): {record.get('grid_cells_mode')}")
    if record.get("trailing_text") not in (None, "", "N/A"):
        lines.append(f"Скользящий стоп-ордер: {record.get('trailing_text')}")
    if record.get("tp_sl_text") not in (None, "", "N/A"):
        lines.append(f"TP/SL: {record.get('tp_sl_text')}")
    if record.get("current_price") is not None:
        lines.append(f"Текущая цена: {format_quote_amount(record.get('current_price'), fallback='N/A')}")
    if record.get("entry_price") is not None:
        lines.append(f"Цена входа: {format_quote_amount(record.get('entry_price'), fallback='N/A')}")
    if record.get("mark_price") is not None:
        lines.append(f"Цена маркировки: {format_quote_amount(record.get('mark_price'), fallback='N/A')}")
    if record.get("liq_price") is not None:
        lines.append(f"Цена ликвидации: {format_quote_amount(record.get('liq_price'), fallback='N/A')}")
    arbitrage_value = record.get("arbitrage_num") or record.get("position_count")
    if arbitrage_value is not None:
        lines.append(f"Кол-во арбитражных сделок: {arbitrage_value}")
    lines.extend([
        f"Причина сигнала: {extra_text}",
        f"ID: {record.get('bot_id') or 'N/A'}"
    ])
    return "\n".join(lines)


def build_risk_alert_markup(alert_type):
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("Отключить этот тип", callback_data=f"alert_risk_disable_{alert_type}"),
        types.InlineKeyboardButton("Отложить на 30 мин", callback_data=f"alert_risk_mute_30m_{alert_type}")
    )
    return markup


def is_risk_alert_muted(alert_type):
    mute_until = safe_float(RISK_ALERT_STATE["mute_until_by_type"].get(alert_type)) or 0.0
    return time.time() < mute_until


def send_admin_notification(message_text, reply_markup=None, parse_mode=None):
    sent = False
    for admin_id in admins:
        try:
            bot.send_message(admin_id, message_text, reply_markup=reply_markup, parse_mode=parse_mode)
            sent = True
        except Exception as e:
            logging.error(f"Ошибка отправки уведомления админу {admin_id}: {e}")
    return sent


def dispatch_active_bot_risk_alerts(active_bots=None):
    if not USE_DB or not admins:
        return 0

    settings = get_notification_settings()
    risk_settings = get_risk_settings()
    active_bots = active_bots if active_bots is not None else fetch_bot_list_data()
    records = [build_active_risk_record(bot_data, index) for index, bot_data in enumerate(active_bots or [])]
    records = [record for record in records if record.get("bot_id")]
    if not records:
        return 0

    symbol_totals = {}
    symbol_direction_counts = {}
    for record in records:
        symbol = record.get("symbol") or "UNKNOWN"
        direction = record.get("direction") or "unknown"
        symbol_totals[symbol] = symbol_totals.get(symbol, 0.0) + (record.get("investment_usdt") or 0.0)
        symbol_direction_counts[(symbol, direction)] = symbol_direction_counts.get((symbol, direction), 0) + 1

    sent_count = 0
    for record in records:
        symbol = record.get("symbol") or "UNKNOWN"
        bot_id = record.get("bot_id")
        bot_type = record.get("bot_type")
        leverage_limit = None
        if bot_type == "GRID_FUTURES":
            leverage_limit = risk_settings.get("max_leverage_grid_futures")
        elif bot_type == "MART_FUTURES":
            leverage_limit = risk_settings.get("max_leverage_mart_futures")

        candidate_alerts = []
        leverage_value = record.get("leverage")
        if (
            settings.get("bot_risk_limit", False)
            and leverage_limit not in (None, 0)
            and leverage_value is not None
            and leverage_value > leverage_limit
        ):
            candidate_alerts.append((
                f"risk:leverage:{bot_id}",
                "bot_risk_limit",
                f"Плечо {leverage_value:.2f}x выше лимита {leverage_limit:.2f}x."
            ))

        total_margin_limit = risk_settings.get("max_total_margin_per_symbol_usdt")
        symbol_total_margin = symbol_totals.get(symbol, 0.0)
        if (
            settings.get("bot_risk_limit", False)
            and total_margin_limit not in (None, 0)
            and symbol_total_margin > total_margin_limit
        ):
            candidate_alerts.append((
                f"risk:symbol_margin:{symbol}",
                "bot_risk_limit",
                f"Суммарная маржа по {symbol}: {format_usdt(symbol_total_margin)} выше лимита {format_usdt(total_margin_limit)}."
            ))

        same_direction_limit = safe_int(risk_settings.get("max_active_bots_same_symbol_direction"))
        same_direction_count = symbol_direction_counts.get((symbol, record.get("direction") or "unknown"), 0)
        if (
            settings.get("bot_risk_limit", False)
            and same_direction_limit not in (None, 0)
            and same_direction_count > same_direction_limit
        ):
            candidate_alerts.append((
                f"risk:symbol_direction:{symbol}:{record.get('direction')}",
                "bot_risk_limit",
                f"Активно {same_direction_count} ботов {record.get('direction')} по {symbol}, лимит {same_direction_limit}."
            ))

        max_loss_alert_pct = risk_settings.get("max_loss_alert_pct")
        pnl_percent = safe_float(record.get("pnl_percent"))
        if (
            settings.get("bot_pnl_drawdown", False)
            and max_loss_alert_pct is not None
            and pnl_percent is not None
            and pnl_percent <= max_loss_alert_pct
        ):
            candidate_alerts.append((
                f"risk:drawdown:{bot_id}",
                "bot_pnl_drawdown",
                f"Текущая просадка {pnl_percent:.2f}% ниже порога {max_loss_alert_pct:.2f}%."
            ))

        mark_price = safe_float(record.get("mark_price"))
        liq_price = safe_float(record.get("liq_price"))
        min_liq_distance_pct = risk_settings.get("min_liq_distance_pct")
        if (
            settings.get("bot_liq_distance", False)
            and min_liq_distance_pct not in (None, 0)
            and mark_price not in (None, 0)
            and liq_price not in (None, 0)
        ):
            liq_distance_pct = abs(mark_price - liq_price) / abs(mark_price) * 100.0
            if liq_distance_pct <= min_liq_distance_pct:
                candidate_alerts.append((
                    f"risk:liq_distance:{bot_id}",
                    "bot_liq_distance",
                    f"До ликвидации осталось около {liq_distance_pct:.2f}% (mark {mark_price}, liq {liq_price})."
                ))

        margin_growth_threshold = risk_settings.get("margin_growth_alert_pct")
        initial_metrics = get_bot_initial_snapshot_metrics(bot_id)
        if (
            settings.get("bot_risk_limit", False)
            and margin_growth_threshold not in (None, 0)
            and initial_metrics
            and initial_metrics.get("investment_usdt") not in (None, 0)
            and record.get("investment_usdt") not in (None, 0)
        ):
            growth_pct = (
                (record["investment_usdt"] - initial_metrics["investment_usdt"])
                / initial_metrics["investment_usdt"] * 100.0
            )
            if growth_pct >= margin_growth_threshold:
                candidate_alerts.append((
                    f"risk:margin_growth:{bot_id}",
                    "bot_risk_limit",
                    f"Маржа выросла на {growth_pct:.2f}% от стартовой {format_usdt(initial_metrics['investment_usdt'])}."
                ))

        recent_loss = find_recent_symbol_loss(symbol, risk_settings.get("repeat_loss_cooldown_hours"))
        if settings.get("bot_repeat_loss", False) and recent_loss:
            recent_end = format_short_datetime(recent_loss.get("ended_ts"), fallback="N/A")
            recent_profit = get_top_bot_profit_value(recent_loss)
            candidate_alerts.append((
                f"risk:repeat_loss:{bot_id}",
                "bot_repeat_loss",
                (
                    f"По {symbol} уже был свежий убыточный бот: "
                    f"{format_usdt(recent_profit, fallback='N/A')} ({recent_end})."
                )
            ))

        for alert_key, alert_type, extra_text in candidate_alerts:
            if is_risk_alert_muted(alert_type):
                continue
            if get_alert_event(alert_key):
                continue
            if send_admin_notification(
                build_risk_alert_message(alert_type, record, extra_text),
                reply_markup=build_risk_alert_markup(alert_type)
            ):
                record_alert_event(alert_key, alert_type, bot_id=bot_id, symbol=symbol, payload=record)
                sent_count += 1
    return sent_count


def get_closed_bots_in_period(start_dt, end_dt):
    if not USE_DB:
        return []
    ensure_db_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT *
        FROM bot_archive
        WHERE COALESCE(is_active, 0) = 0
          AND COALESCE(ended_ts, created_ts, 0) BETWEEN ? AND ?
        ORDER BY COALESCE(ended_ts, created_ts, 0) ASC
        """,
        (int(start_dt.timestamp()), int(end_dt.timestamp()))
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows


def build_closed_bots_report(period_name, start_dt, end_dt):
    rows = get_closed_bots_in_period(start_dt, end_dt)
    if not rows:
        return f"Отчёт {period_name}\nНет закрытых ботов за период."

    profits = [get_top_bot_profit_value(row) or 0.0 for row in rows]
    total_profit = sum(profits)
    winners = [value for value in profits if value > 0]
    losers = [value for value in profits if value < 0]
    sorted_rows = sorted(rows, key=lambda row: get_top_bot_profit_value(row) or 0.0)
    worst_rows = sorted_rows[:3]
    best_rows = list(reversed(sorted_rows[-3:]))

    close_groups = {}
    for row in rows:
        key = row.get("close_code") or row.get("close_reason") or "UNKNOWN"
        close_groups[key] = close_groups.get(key, 0) + 1

    lines = [
        f"Отчёт {period_name}",
        f"Период: {start_dt.strftime('%d.%m.%Y %H:%M')} - {end_dt.strftime('%d.%m.%Y %H:%M')}",
        f"Закрыто ботов: {len(rows)}",
        f"Итоговый P&L: {format_usdt(total_profit)}",
        f"Плюсовых: {len(winners)} | Минусовых: {len(losers)}",
        ""
    ]
    lines.append("Причины закрытия:")
    for key, count in sorted(close_groups.items(), key=lambda item: item[1], reverse=True)[:5]:
        lines.append(f"{key}: {count}")
    lines.append("")
    lines.append("Худшие:")
    for row in worst_rows:
        lines.append(f"{row.get('symbol')}: {format_usdt(get_top_bot_profit_value(row), fallback='N/A')}")
    lines.append("")
    lines.append("Лучшие:")
    for row in best_rows:
        lines.append(f"{row.get('symbol')}: {format_usdt(get_top_bot_profit_value(row), fallback='N/A')}")
    return "\n".join(lines[:25])


def get_latest_balance_breakdown_row():
    if not USE_DB:
        return None
    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT date, non_bot_balance
        FROM balances
        WHERE non_bot_balance IS NOT NULL
        ORDER BY date DESC
        LIMIT 1
        """
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    return {
        "date": row[0],
        "non_bot_balance": safe_float(row[1])
    }


def get_snapshot_rows_at_or_before(cursor, target_time):
    cursor.execute(
        "SELECT MAX(snapshot_time) FROM bot_snapshots WHERE snapshot_time <= ?",
        (target_time,)
    )
    snapshot_time_row = cursor.fetchone()
    snapshot_time = snapshot_time_row[0] if snapshot_time_row else None
    if not snapshot_time:
        return []
    cursor.execute(
        """
        SELECT bot_id, symbol, bot_type, investment_usdt
        FROM bot_snapshots
        WHERE snapshot_time = ?
        """,
        (snapshot_time,)
    )
    rows = []
    for bot_id, symbol, bot_type, investment_usdt in cursor.fetchall():
        rows.append({
            "bot_id": bot_id,
            "symbol": symbol,
            "bot_type": bot_type,
            "investment_usdt": safe_float(investment_usdt),
            "identity_key": get_bot_identity_key(bot_id, symbol, bot_type)
        })
    return rows


def match_duplicate_non_bot_jump(extra_non_bot, current_records, previous_records):
    previous_keys = {record.get("identity_key") for record in previous_records}
    for record in current_records:
        if record.get("identity_key") in previous_keys:
            continue
        invested = safe_float(record.get("investment_usdt"))
        if invested is None or invested < BOT_DUPLICATE_MIN_JUMP_USDT:
            continue
        tolerance = max(BOT_DUPLICATE_MATCH_ABS_USDT, abs(invested) * BOT_DUPLICATE_MATCH_RATIO)
        if abs(extra_non_bot - invested) <= tolerance:
            return record, invested
    return None, None


def correct_duplicate_non_bot_balance(non_bot_balance, current_records):
    current_non_bot = safe_float(non_bot_balance)
    if current_non_bot is None:
        return non_bot_balance, None

    previous_balance_row = get_latest_balance_breakdown_row()
    if previous_balance_row is None:
        return current_non_bot, None

    previous_non_bot = previous_balance_row.get("non_bot_balance") or 0.0
    extra_non_bot = current_non_bot - previous_non_bot
    if extra_non_bot < BOT_DUPLICATE_MIN_JUMP_USDT:
        return current_non_bot, None

    conn = get_db_connection()
    cursor = conn.cursor()
    previous_snapshot_rows = get_snapshot_rows_at_or_before(cursor, previous_balance_row["date"])
    conn.close()

    duplicate_record, matched_investment = match_duplicate_non_bot_jump(
        extra_non_bot,
        current_records,
        previous_snapshot_rows
    )
    if duplicate_record is None:
        return current_non_bot, None

    corrected_non_bot = max(0.0, current_non_bot - matched_investment)
    return corrected_non_bot, duplicate_record


def repair_duplicate_bot_balance_spikes(limit_rows=1440):
    if not USE_DB:
        return 0

    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    if limit_rows is None:
        cursor.execute(
            """
            SELECT date, current_balance, balance_in_usd, balance_rub, bot_balance, funding_balance, non_bot_balance
            FROM balances
            ORDER BY date ASC
            """
        )
    else:
        cursor.execute(
            """
            SELECT date, current_balance, balance_in_usd, balance_rub, bot_balance, funding_balance, non_bot_balance
            FROM (
                SELECT date, current_balance, balance_in_usd, balance_rub, bot_balance, funding_balance, non_bot_balance
                FROM balances
                ORDER BY date DESC
                LIMIT ?
            )
            ORDER BY date ASC
            """,
            (int(limit_rows),)
        )
    rows = cursor.fetchall()
    repaired_rows = 0

    for idx in range(1, len(rows)):
        date_str, current_balance, balance_in_usd, balance_rub, bot_balance, funding_balance, non_bot_balance = rows[idx]
        previous_non_bot = safe_float(rows[idx - 1][6]) or 0.0
        current_non_bot = safe_float(non_bot_balance)
        current_bot_balance = safe_float(bot_balance)
        current_total = safe_float(balance_in_usd) or safe_float(current_balance)
        current_rub = safe_float(balance_rub)
        if current_non_bot is None or current_bot_balance is None or current_total is None:
            continue

        extra_non_bot = current_non_bot - previous_non_bot
        if extra_non_bot < BOT_DUPLICATE_MIN_JUMP_USDT:
            continue

        previous_snapshot_rows = get_snapshot_rows_at_or_before(cursor, rows[idx - 1][0])
        current_snapshot_rows = get_snapshot_rows_at_or_before(cursor, date_str)
        duplicate_record, matched_investment = match_duplicate_non_bot_jump(
            extra_non_bot,
            current_snapshot_rows,
            previous_snapshot_rows
        )
        if duplicate_record is None:
            continue

        corrected_non_bot = max(0.0, current_non_bot - matched_investment)
        corrected_total = current_bot_balance + corrected_non_bot
        rub_ratio = current_rub / current_total if current_rub not in (None, 0) and current_total not in (None, 0) else None
        corrected_rub = corrected_total * rub_ratio if rub_ratio is not None else current_rub

        cursor.execute(
            """
            UPDATE balances
            SET current_balance = ?, balance_in_usd = ?, balance_rub = COALESCE(?, balance_rub),
                funding_balance = ?, non_bot_balance = ?
            WHERE date = ?
            """,
            (
                corrected_total,
                corrected_total,
                corrected_rub,
                corrected_non_bot,
                corrected_non_bot,
                date_str
            )
        )
        repaired_rows += 1

    conn.commit()
    conn.close()
    return repaired_rows


def format_decimal(value, digits=2, fallback="N/A"):
    number = safe_float(value)
    if number is None:
        return fallback if value in (None, "") else str(value)
    return f"{number:,.{digits}f}".replace(",", " ")


def format_usdt(value, digits=2, fallback="N/A"):
    number = safe_float(value)
    if number is None:
        return fallback if value in (None, "") else f"{value} USDT"
    return f"{format_decimal(number, digits)} USDT"


def format_quote_amount(value, digits=None, fallback="N/A"):
    number = safe_float(value)
    if number is None:
        return fallback if value in (None, "") else str(value)
    if digits is None:
        abs_number = abs(number)
        if abs_number >= 100:
            digits = 2
        elif abs_number >= 1:
            digits = 4
        elif abs_number >= 0.01:
            digits = 6
        else:
            digits = 8
    return format_decimal(number, digits)


def format_percent(value, scale=1.0, digits=2, fallback="N/A"):
    number = safe_float(value)
    if number is None:
        return fallback if value in (None, "") else str(value)
    return f"{number * scale:.{digits}f}%"


def short_text(value, limit=28):
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return f"{text[:max(1, limit - 1)]}…"


def date_to_token(value):
    return value.strftime("%Y%m%d")


def token_to_date(token):
    return datetime.strptime(token, "%Y%m%d").date()


def cleanup_graph_cache(max_dirs=24, min_interval_seconds=3600):
    os.makedirs(GRAPH_DIR, exist_ok=True)
    now_ts = time.time()
    if now_ts - GRAPH_CACHE_STATE["last_cleanup_ts"] < min_interval_seconds:
        return

    GRAPH_CACHE_STATE["last_cleanup_ts"] = now_ts
    cache_dirs = []
    for name in os.listdir(GRAPH_DIR):
        folder_path = os.path.join(GRAPH_DIR, name)
        if not os.path.isdir(folder_path):
            continue
        try:
            sort_key = datetime.strptime(name, "%Y%m%d")
        except ValueError:
            sort_key = datetime.fromtimestamp(os.path.getmtime(folder_path))
        cache_dirs.append((sort_key, folder_path))

    for _, folder_path in sorted(cache_dirs, key=lambda item: item[0], reverse=True)[max_dirs:]:
        shutil.rmtree(folder_path, ignore_errors=True)


def get_graph_cache_dir(selected_date):
    cleanup_graph_cache()
    folder = os.path.join(GRAPH_DIR, date_to_token(selected_date))
    os.makedirs(folder, exist_ok=True)
    return folder


def get_overview_graph_path(selected_date):
    return os.path.join(get_graph_cache_dir(selected_date), "overview.png")


def get_bot_card_path(selected_date, bot_index):
    return os.path.join(get_graph_cache_dir(selected_date), f"bot_{bot_index:02d}.png")


def get_bot_type_name(bot_type):
    return {
        "GRID_FUTURES": "Фьючерсный grid-бот",
        "MART_FUTURES": "Фьючерсный Мартингейл",
        "GRID_SPOT": "Спотовый grid-бот",
        "COMBO_FUTURES": "Фьючерсный combo-бот"
    }.get(bot_type, "Bybit бот")


def get_bot_button_kind(bot_type):
    return {
        "GRID_FUTURES": "Grid",
        "MART_FUTURES": "Mart",
        "GRID_SPOT": "Spot",
        "COMBO_FUTURES": "Combo"
    }.get(bot_type, "Bot")


def get_card_palette(bot_type, pnl_value):
    pnl_number = safe_float(pnl_value)
    if pnl_number is not None and pnl_number < 0:
        return {
            "accent": "#b45309",
            "header": "#fff7ed",
            "surface": "#ffffff"
        }
    if bot_type == "MART_FUTURES":
        return {
            "accent": "#1d4ed8",
            "header": "#eff6ff",
            "surface": "#ffffff"
        }
    if bot_type == "GRID_SPOT":
        return {
            "accent": "#0f766e",
            "header": "#ecfeff",
            "surface": "#ffffff"
        }
    return {
        "accent": "#7c3aed",
        "header": "#f5f3ff",
        "surface": "#ffffff"
    }


def format_full_datetime(ts_value, fallback="N/A"):
    normalized = normalize_epoch_timestamp(ts_value)
    if normalized is None:
        return fallback
    return datetime.fromtimestamp(normalized).strftime('%Y-%m-%d %H:%M:%S')


def derive_start_time_text(start_ts=None, running_duration=None, end_ts=None, fallback="N/A"):
    normalized_start = normalize_epoch_timestamp(start_ts)
    if normalized_start is not None:
        return datetime.fromtimestamp(normalized_start).strftime('%Y-%m-%d %H:%M:%S')
    duration_seconds = safe_int(running_duration)
    if duration_seconds is None:
        return fallback
    end_value = normalize_epoch_timestamp(end_ts) or int(time.time())
    if end_value < duration_seconds:
        return fallback
    return datetime.fromtimestamp(end_value - duration_seconds).strftime('%Y-%m-%d %H:%M:%S')


def format_percent_auto(value, digits=2, fallback="N/A"):
    number = safe_float(value)
    if number is None:
        return fallback
    scale = 100 if abs(number) <= 1 else 1
    return format_percent(number, scale=scale, digits=digits, fallback=fallback)


def format_bot_status(status_value, fallback="N/A"):
    raw_value = str(status_value or "").strip()
    if not raw_value:
        return fallback
    upper_value = raw_value.upper()
    mapping = {
        "RUNNING": "Активный",
        "ACTIVE": "Активный",
        "COMPLETED": "Закрыт",
        "CLOSED": "Закрыт",
        "F_MART_BOT_DISPLAY_STATUS_COMPLETED": "Закрыт",
        "F_MART_BOT_DISPLAY_STATUS_RUNNING": "Активный",
        "BOT_STATUS_RUNNING": "Активный",
        "BOT_STATUS_COMPLETED": "Закрыт"
    }
    return mapping.get(upper_value, raw_value.replace("_", " ").title())


def format_grid_distribution(value, fallback="N/A"):
    raw_value = str(value or "").strip()
    if not raw_value:
        return fallback
    upper_value = raw_value.upper()
    if "ARITH" in upper_value:
        return "Арифметическое"
    if "GEOM" in upper_value:
        return "Геометрическое"
    return raw_value.replace("_", " ").title()


def build_grid_cells_mode_text(cell_count, grid_type=None, fallback="N/A"):
    count_value = safe_int(cell_count)
    type_label = format_grid_distribution(grid_type, fallback="")
    if count_value is None and not type_label:
        return fallback
    if count_value is None:
        return type_label or fallback
    if type_label:
        return f"{count_value} ({type_label})"
    return str(count_value)


def build_tp_sl_text(tp_value=None, sl_value=None):
    tp_text = format_percent_auto(tp_value, fallback="--")
    sl_text = format_percent_auto(sl_value, fallback="--")
    if tp_text == "--" and sl_text == "--":
        return "N/A"
    return f"{tp_text}/{sl_text}"


def build_price_pair_text(first_value=None, second_value=None):
    first_text = format_quote_amount(first_value, fallback="--")
    second_text = format_quote_amount(second_value, fallback="--")
    if first_text == "--" and second_text == "--":
        return "N/A"
    return f"{first_text}/{second_text}"


def build_trailing_text(percent_value=None, enabled=None):
    percent_text = format_percent_auto(percent_value, fallback="N/A")
    if enabled is None:
        return percent_text
    enabled_text = "вкл" if bool(enabled) else "выкл"
    if percent_text == "N/A":
        return enabled_text
    return f"{percent_text} ({enabled_text})"


def build_bot_snapshot(bot_data, index):
    base_snapshot = {
        "index": index,
        "symbol": None,
        "bot_type": "UNKNOWN",
        "title": f"Бот {index + 1}",
        "button_label": f"Бот {index + 1}",
        "subtitle": "Нет данных о боте",
        "badge": "N/A",
        "palette": {
            "accent": "#475569",
            "header": "#f8fafc",
            "surface": "#ffffff"
        },
        "metrics": [("Статус", "Нет данных")],
        "details": [("Описание", "Нет данных о боте.")],
        "overview_lines": [("Статус", "Нет данных")],
        "caption_title": f"Бот {index + 1}",
        "investment_usdt": None,
        "pnl_usdt": None,
        "equity_usdt": None,
        "pnl_percent_value": None,
        "status_text": "N/A",
        "started_at_text": "N/A",
        "ended_at_text": "N/A",
        "grid_cells_mode": "N/A",
        "mark_price": None,
        "liq_price": None,
        "current_price": None,
        "entry_price": None,
        "realized_pnl_usdt": None,
        "unrealized_pnl_usdt": None,
        "arbitrage_num": None,
        "trailing_text": "N/A",
        "tp_sl_text": "N/A",
        "source_note": "Текущий снимок Bybit",
        "footer_left": "Источник: Bybit",
        "highlight_primary_label": "Доход",
        "highlight_secondary_label": "Доход %",
        "highlight_primary_value": None,
        "highlight_secondary_value": None,
        "detail_titles": ("Параметры", "Детали")
    }
    if not isinstance(bot_data, dict):
        return base_snapshot

    bot_type, future_grid, futures_mart, spot_grid, spot_profit, combo = get_bot_detail_payload(bot_data)

    symbol = "N/A"
    badge = get_bot_button_kind(bot_type)
    details = []
    metrics = []
    overview_lines = []
    pnl_value = None
    invested = None
    pnl_percent_value = None
    status_text = "N/A"
    started_at_text = "N/A"
    ended_at_text = "N/A"
    grid_cells_mode = "N/A"
    mark_price = None
    liq_price = None
    current_price = None
    entry_price = None
    realized_pnl_value = None
    unrealized_pnl_value = None
    arbitrage_num = None
    trailing_text = "N/A"
    tp_sl_text = "N/A"
    source_note = "Текущий снимок Bybit"
    footer_left = None

    if bot_type == "GRID_FUTURES" and future_grid:
        symbol = future_grid.get("symbol", "N/A")
        mode = future_grid.get("grid_mode", "").lower()
        leverage = future_grid.get("leverage", "N/A")
        if "neutral" in mode:
            badge = f"Neutral {leverage}x"
        elif "long" in mode:
            badge = f"Long {leverage}x"
        elif "short" in mode:
            badge = f"Short {leverage}x"
        invested = future_grid.get("total_investment", "N/A")
        pnl_value = future_grid.get("pnl", future_grid.get("realized_pnl", 0))
        raw_pnl_percent = safe_float(future_grid.get("pnl_per", 0))
        pnl_percent_value = raw_pnl_percent * 100 if raw_pnl_percent is not None else None
        pnl_percent = format_percent(future_grid.get("pnl_per", 0), scale=100)
        status_text = format_bot_status(future_grid.get("status"))
        started_at_text = derive_start_time_text(
            future_grid.get("create_time"),
            future_grid.get("running_duration"),
            future_grid.get("end_time")
        )
        ended_at_text = format_full_datetime(future_grid.get("end_time"), fallback="Активен")
        grid_cells_mode = build_grid_cells_mode_text(future_grid.get("cell_num"), future_grid.get("grid_type"))
        mark_price = safe_float(future_grid.get("mark_price"))
        liq_price = safe_float(future_grid.get("liq_price"))
        current_price = safe_float(future_grid.get("current_price"))
        entry_price = safe_float(future_grid.get("entry_price"))
        if entry_price == 0:
            entry_price = None
        arbitrage_num = safe_int(future_grid.get("arbitrage_num"))
        runtime = format_duration(
            future_grid.get("running_duration", "0")
        ) if safe_int(future_grid.get("running_duration")) is not None else format_duration_from_timestamps(
            future_grid.get("create_time"),
            future_grid.get("end_time")
        )
        price_range = f"{future_grid.get('min_price', 'N/A')} - {future_grid.get('max_price', 'N/A')}"
        metrics = [
            ("Инвестиции", format_usdt(invested)),
            ("Общий P&L", format_usdt(pnl_value, fallback="N/A")),
            ("% P&L", pnl_percent),
            ("Арб. сделки", str(arbitrage_num) if arbitrage_num is not None else "N/A")
        ]
        details = [
            ("Время запуска", started_at_text),
            ("Статус", status_text),
            ("Ценовой диапазон", price_range),
            ("Сетки (режим)", grid_cells_mode),
            ("Текущая цена", format_quote_amount(current_price, fallback="N/A")),
            ("Цена входа", format_quote_amount(entry_price, fallback="N/A")),
            ("Цена маркировки", format_quote_amount(mark_price, fallback="N/A")),
            ("Цена ликвидации", format_quote_amount(liq_price, fallback="N/A")),
            ("APR", format_percent_auto(future_grid.get("total_apr"), fallback="N/A")),
            ("Активен", runtime)
        ]
        close_detail = future_grid.get("close_detail") or {}
        close_reason = close_detail.get("close_reason") or close_detail.get("bot_close_code")
        if close_reason:
            details.append(("Причина закрытия", str(close_reason)))
        overview_lines = [
            ("Режим", badge),
            ("Статус", status_text),
            ("P&L", f"{format_usdt(pnl_value, fallback='N/A')} / {pnl_percent}"),
            ("Инвестиции", format_usdt(invested)),
            ("Диапазон", short_text(price_range, 26)),
            ("Сетки", grid_cells_mode)
        ]
    elif bot_type == "MART_FUTURES" and futures_mart:
        symbol = futures_mart.get("symbol", "N/A")
        mode = futures_mart.get("fmart_mode", "").lower()
        leverage = futures_mart.get("leverage", "N/A")
        if "neutral" in mode:
            badge = f"Neutral {leverage}x"
        elif "long" in mode:
            badge = f"Long {leverage}x"
        elif "short" in mode:
            badge = f"Short {leverage}x"
        invested = futures_mart.get("total_margin", futures_mart.get("total_investment", "N/A"))
        realized_pnl_value = safe_float(futures_mart.get("realized_pnl"))
        unrealized_pnl_value = safe_float(futures_mart.get("unrealized_pnl"))
        pnl_value = futures_mart.get("total_profit", futures_mart.get("realized_pnl", 0))
        raw_pnl_percent = safe_float(futures_mart.get("total_profit_per"))
        invested_value = safe_float(invested)
        pnl_numeric = safe_float(pnl_value)
        status_text = format_bot_status(futures_mart.get("bot_display_status") or futures_mart.get("status"))
        started_at_text = derive_start_time_text(
            futures_mart.get("create_time"),
            futures_mart.get("running_duration"),
            futures_mart.get("end_time")
        )
        ended_at_text = format_full_datetime(futures_mart.get("end_time"), fallback="Активен")
        mark_price = safe_float(futures_mart.get("current_price"))
        liq_price = safe_float(futures_mart.get("liq_price"))
        entry_price = safe_float(futures_mart.get("entry_price"))
        current_price = safe_float(futures_mart.get("current_price"))
        if entry_price == 0:
            entry_price = None
        if liq_price == 0:
            liq_price = None
        tp_sl_text = build_tp_sl_text(futures_mart.get("round_tp_percent"), futures_mart.get("sl_per"))
        if raw_pnl_percent is not None:
            pnl_percent_value = raw_pnl_percent * 100
            pnl_percent = format_percent(raw_pnl_percent, scale=100)
        elif invested_value not in (None, 0) and pnl_numeric is not None:
            pnl_percent_value = pnl_numeric / invested_value * 100
            pnl_percent = format_percent(pnl_percent_value, scale=1)
        else:
            pnl_percent_value = None
            pnl_percent = "N/A"
        runtime = format_duration(
            futures_mart.get("running_duration", "0")
        ) if safe_int(futures_mart.get("running_duration")) is not None else format_duration_from_timestamps(
            futures_mart.get("create_time"),
            futures_mart.get("end_time")
        )
        price_drop = format_percent(futures_mart.get("price_float_per", 0), scale=100, digits=1)
        metrics = [
            ("Маржа", format_usdt(invested)),
            ("Общий P&L", format_usdt(pnl_value, fallback="N/A")),
            ("Реал. P&L", format_usdt(realized_pnl_value, fallback="N/A")),
            ("Нереал. P&L", format_usdt(unrealized_pnl_value, fallback="N/A"))
        ]
        details = [
            ("Время запуска", started_at_text),
            ("Статус", status_text),
            ("TP/SL", tp_sl_text),
            ("Снижение цены", price_drop),
            ("Множитель позиции", futures_mart.get("add_pos_per", "N/A")),
            ("Доборов", futures_mart.get("added_pos_num", futures_mart.get("add_pos_num", "N/A"))),
            ("Текущая цена", format_quote_amount(current_price, fallback="N/A")),
            ("Цена входа", format_quote_amount(entry_price, fallback="N/A")),
            ("Цена ликвидации", format_quote_amount(liq_price, fallback="N/A")),
            ("Остановлен", format_bot_status(futures_mart.get("stop_type"), fallback="N/A"))
        ]
        overview_lines = [
            ("Режим", badge),
            ("Статус", status_text),
            ("P&L", f"{format_usdt(pnl_value, fallback='N/A')} / {pnl_percent}"),
            ("Маржа", format_usdt(invested)),
            ("TP/SL", tp_sl_text),
            ("Снижение", price_drop)
        ]
    elif bot_type == "GRID_SPOT" and spot_grid:
        symbol = spot_grid.get("symbol", "N/A")
        mode = spot_grid.get("grid_mode", "").lower()
        if "neutral" in mode:
            badge = "Neutral"
        elif "long" in mode:
            badge = "Long"
        elif "short" in mode:
            badge = "Short"
        invested = spot_grid.get("total_investment", "N/A")
        pnl_value = spot_profit.get("total_profit", 0)
        realized_pnl_value = safe_float(spot_profit.get("grid_profit", spot_profit.get("total_profit")))
        arbitrage_num = safe_int(spot_profit.get("arbitrage_num"))
        status_text = format_bot_status(spot_grid.get("status"))
        started_at_text = derive_start_time_text(
            spot_grid.get("create_time"),
            spot_grid.get("running_duration"),
            spot_grid.get("modify_time")
        )
        ended_at_text = format_full_datetime(spot_grid.get("modify_time"), fallback="Активен")
        current_price = safe_float(spot_grid.get("current_price"))
        entry_price = safe_float(spot_grid.get("entry_price"))
        if entry_price == 0:
            entry_price = None
        tp_sl_text = build_price_pair_text(spot_grid.get("take_profit_price"), spot_grid.get("stop_loss_price"))
        trailing_text = build_trailing_text(spot_grid.get("ts_percent"), spot_grid.get("enable_trailing_up"))
        grid_cells_mode = build_grid_cells_mode_text(spot_grid.get("cell_number"), spot_grid.get("grid_type"))
        raw_apr = safe_float(spot_profit.get("total_apr"))
        invested_value = safe_float(invested)
        pnl_numeric = safe_float(pnl_value)
        if raw_apr is not None:
            pnl_percent_value = raw_apr * 100
            pnl_percent = format_percent(raw_apr, scale=100)
        elif invested_value not in (None, 0) and pnl_numeric is not None:
            pnl_percent_value = pnl_numeric / invested_value * 100
            pnl_percent = format_percent(pnl_percent_value, scale=1)
        else:
            pnl_percent_value = None
            pnl_percent = "N/A"
        runtime = format_duration(
            spot_grid.get("running_duration", "0")
        ) if safe_int(spot_grid.get("running_duration")) is not None else format_duration_from_timestamps(
            spot_grid.get("create_time"),
            spot_grid.get("modify_time")
        )
        price_range = f"{spot_grid.get('min_price', 'N/A')} - {spot_grid.get('max_price', 'N/A')}"
        metrics = [
            ("Инвестиции", format_usdt(invested)),
            ("Общий P&L", format_usdt(pnl_value, fallback="N/A")),
            ("Получ. прибыль", format_usdt(realized_pnl_value, fallback="N/A")),
            ("Арб. сделки", str(arbitrage_num) if arbitrage_num is not None else "N/A")
        ]
        details = [
            ("Время запуска", started_at_text),
            ("Статус", status_text),
            ("Ценовой диапазон", price_range),
            ("Сетки (режим)", grid_cells_mode),
            ("Текущая цена", format_quote_amount(current_price, fallback="N/A")),
            ("Цена входа", format_quote_amount(entry_price, fallback="N/A")),
            ("Трейлинг", trailing_text),
            ("TP/SL", tp_sl_text),
            ("APR", pnl_percent),
            ("Активен", runtime)
        ]
        overview_lines = [
            ("Режим", badge),
            ("Статус", status_text),
            ("P&L", f"{format_usdt(pnl_value, fallback='N/A')} / {pnl_percent}"),
            ("Инвестиции", format_usdt(invested)),
            ("Диапазон", short_text(price_range, 26)),
            ("Сетки", grid_cells_mode)
        ]
    elif bot_type == "COMBO_FUTURES" and combo:
        symbol_settings = combo.get("symbol_settings") or []
        symbol_list = []
        for item in symbol_settings:
            token = (
                (item or {}).get("symbol")
                or (item or {}).get("base_token")
                or (item or {}).get("coin")
            )
            if token:
                symbol_list.append(str(token))
        symbol = " + ".join(symbol_list[:3]) if symbol_list else combo.get("symbol", "N/A")
        mode = str(combo.get("bot_mode", "")).lower()
        leverage = combo.get("leverage", "N/A")
        if "neutral" in mode:
            badge = f"Neutral {leverage}x"
        elif "long" in mode:
            badge = f"Long {leverage}x"
        elif "short" in mode:
            badge = f"Short {leverage}x"
        invested = combo.get("total_margin", combo.get("margin", "N/A"))
        realized_pnl_value = safe_float(combo.get("realized_pnl"))
        unrealized_pnl_value = safe_float(combo.get("unrealized_pnl"))
        pnl_value = combo.get("total_pnl", combo.get("realized_pnl", 0))
        raw_pnl_percent = safe_float(combo.get("total_pnl_per"))
        invested_value = safe_float(invested)
        pnl_numeric = safe_float(pnl_value)
        status_text = format_bot_status(combo.get("status") or combo.get("bot_display_status"))
        started_at_text = derive_start_time_text(
            combo.get("create_time"),
            combo.get("run_time_duration"),
            combo.get("end_time")
        )
        ended_at_text = format_full_datetime(combo.get("end_time"), fallback="Активен")
        mark_price = safe_float(combo.get("mark_price"))
        liq_price = safe_float(combo.get("liq_price"))
        if raw_pnl_percent is not None:
            pnl_percent_value = raw_pnl_percent * 100
            pnl_percent = format_percent(raw_pnl_percent, scale=100)
        elif invested_value not in (None, 0) and pnl_numeric is not None:
            pnl_percent_value = pnl_numeric / invested_value * 100
            pnl_percent = format_percent(pnl_percent_value, scale=1)
        else:
            pnl_percent_value = None
            pnl_percent = "N/A"
        runtime = format_duration(
            combo.get("run_time_duration", "0")
        ) if safe_int(combo.get("run_time_duration")) is not None else format_duration_from_timestamps(
            combo.get("create_time"),
            combo.get("end_time")
        )
        metrics = [
            ("Маржа", format_usdt(invested)),
            ("Общий P&L", format_usdt(pnl_value, fallback="N/A")),
            ("Реал. P&L", format_usdt(realized_pnl_value, fallback="N/A")),
            ("Нереал. P&L", format_usdt(unrealized_pnl_value, fallback="N/A"))
        ]
        details = [
            ("Время запуска", started_at_text),
            ("Статус", status_text),
            ("Символы", short_text(", ".join(symbol_list) if symbol_list else symbol, 18)),
            ("Ног", combo.get("leg_num", len(symbol_list) or "N/A")),
            ("Позиции", combo.get("position_num", combo.get("symbol_count", "N/A"))),
            ("Цена маркировки", format_quote_amount(mark_price, fallback="N/A")),
            ("Цена ликвидации", format_quote_amount(liq_price, fallback="N/A")),
            ("Активен", runtime)
        ]
        overview_lines = [
            ("Режим", badge),
            ("Статус", status_text),
            ("P&L", f"{format_usdt(pnl_value, fallback='N/A')} / {pnl_percent}"),
            ("Маржа", format_usdt(invested)),
            ("Символы", short_text(symbol, 24)),
            ("Позиции", combo.get("position_num", combo.get("symbol_count", "N/A")))
        ]

    if symbol == "N/A":
        title = f"Бот {index + 1}"
    else:
        title = symbol

    palette = get_card_palette(bot_type, pnl_value)
    invested_value = safe_float(invested)
    pnl_numeric = safe_float(pnl_value)
    equity_value = None
    if invested_value is not None and pnl_numeric is not None:
        equity_value = invested_value + pnl_numeric
    elif invested_value is not None:
        equity_value = invested_value

    button_label = short_text(f"{title} · {get_bot_button_kind(bot_type)}", 24)
    highlight_primary_value = format_usdt(pnl_numeric, fallback="N/A")
    highlight_secondary_value = format_percent(pnl_percent_value, scale=1, fallback="N/A")
    footer_left = footer_left or f"Источник: Bybit • {get_bot_type_name(bot_type)}"
    return {
        "index": index,
        "symbol": None if symbol == "N/A" else symbol,
        "bot_type": bot_type,
        "title": title,
        "button_label": button_label,
        "subtitle": get_bot_type_name(bot_type),
        "badge": badge,
        "palette": palette,
        "metrics": metrics or [("Статус", "Нет данных")],
        "details": details or [("Описание", "Нет данных о боте.")],
        "overview_lines": overview_lines or [("Статус", "Нет данных")],
        "caption_title": title,
        "investment_usdt": invested_value,
        "pnl_usdt": pnl_numeric,
        "equity_usdt": equity_value,
        "pnl_percent_value": pnl_percent_value,
        "status_text": status_text,
        "started_at_text": started_at_text,
        "ended_at_text": ended_at_text,
        "grid_cells_mode": grid_cells_mode,
        "mark_price": mark_price,
        "liq_price": liq_price,
        "current_price": current_price,
        "entry_price": entry_price,
        "realized_pnl_usdt": realized_pnl_value,
        "unrealized_pnl_usdt": unrealized_pnl_value,
        "arbitrage_num": arbitrage_num,
        "trailing_text": trailing_text,
        "tp_sl_text": tp_sl_text,
        "source_note": source_note,
        "footer_left": footer_left,
        "highlight_primary_label": "Доход",
        "highlight_secondary_label": "Доход %",
        "highlight_primary_value": highlight_primary_value,
        "highlight_secondary_value": highlight_secondary_value,
        "detail_titles": ("Параметры", "Детали")
    }


def persist_bot_snapshots(snapshot_time, bots_data):
    if not USE_DB or not bots_data:
        return

    ensure_db_schema()
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        archive_records = build_bot_archive_records(bots_data, is_active=True)
        for index, bot_data in enumerate(bots_data):
            snapshot = build_bot_snapshot(bot_data, index)
            archive_record = archive_records[index] if index < len(archive_records) else {}
            if (
                snapshot["symbol"] is None
                and snapshot["investment_usdt"] is None
                and snapshot["pnl_usdt"] is None
            ):
                continue

            cursor.execute(
                """
                REPLACE INTO bot_snapshots (
                    snapshot_time, bot_index, bot_id, symbol, bot_type, title, badge,
                    investment_usdt, pnl_usdt, equity_usdt, pnl_percent,
                    status, display_status, is_active, close_code
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_time,
                    snapshot["index"],
                    archive_record.get("bot_id"),
                    snapshot["symbol"],
                    snapshot["bot_type"],
                    snapshot["title"],
                    snapshot["badge"],
                    snapshot["investment_usdt"],
                    snapshot["pnl_usdt"],
                    snapshot["equity_usdt"],
                    snapshot["pnl_percent_value"],
                    archive_record.get("status"),
                    archive_record.get("display_status"),
                    archive_record.get("is_active"),
                    archive_record.get("close_code")
                )
            )
        conn.commit()
    finally:
        conn.close()


def get_bot_day_history(selected_date, snapshot):
    if not USE_DB:
        return []

    ensure_db_schema()
    day_prefix = selected_date.strftime('%Y-%m-%d')
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        rows = []
        symbol = snapshot.get("symbol")
        bot_type = snapshot.get("bot_type")
        if symbol:
            cursor.execute(
                """
                SELECT snapshot_time, investment_usdt, pnl_usdt, equity_usdt
                FROM bot_snapshots
                WHERE snapshot_time LIKE ? AND symbol = ? AND bot_type = ?
                ORDER BY snapshot_time ASC
                """,
                (f"{day_prefix}%", symbol, bot_type)
            )
            rows = cursor.fetchall()

        if not rows:
            cursor.execute(
                """
                SELECT snapshot_time, investment_usdt, pnl_usdt, equity_usdt
                FROM bot_snapshots
                WHERE snapshot_time LIKE ? AND bot_index = ?
                ORDER BY snapshot_time ASC
                """,
                (f"{day_prefix}%", snapshot["index"])
            )
            rows = cursor.fetchall()
    finally:
        conn.close()

    history = []
    for snapshot_time, investment_usdt, pnl_usdt, equity_usdt in rows:
        try:
            point_time = datetime.strptime(snapshot_time, '%Y-%m-%d %H:%M:%S')
        except Exception:
            continue
        history.append({
            "time": point_time,
            "investment_usdt": safe_float(investment_usdt),
            "pnl_usdt": safe_float(pnl_usdt),
            "equity_usdt": safe_float(equity_usdt)
        })
    return history


def draw_summary_box(ax, title, values, anchor=(0.985, 0.95)):
    if not values:
        return
    text = (
        f"{title}\n"
        f"Max: {format_decimal(max(values))}\n"
        f"Avg: {format_decimal(sum(values) / len(values))}\n"
        f"Min: {format_decimal(min(values))}"
    )
    ax.text(
        anchor[0],
        anchor[1],
        text,
        transform=ax.transAxes,
        ha='right',
        va='top',
        fontsize=7.5,
        fontfamily='DejaVu Sans Mono',
        bbox={
            "boxstyle": "round,pad=0.35",
            "facecolor": "#ffffff",
            "edgecolor": "#cbd5e1",
            "linewidth": 1
        }
    )


def get_day_annotation_indices(values):
    if not values:
        return []

    indices = {0, len(values) - 1}
    if len(values) > 2:
        max_idx = max(range(len(values)), key=lambda idx: values[idx])
        min_idx = min(range(len(values)), key=lambda idx: values[idx])
        indices.update({max_idx, min_idx})

    if len(values) > 6:
        span = max(values) - min(values)
        threshold = span * 0.18 if span > 0 else 1
        for idx in range(1, len(values) - 1):
            left_delta = values[idx] - values[idx - 1]
            right_delta = values[idx + 1] - values[idx]
            is_turn = left_delta == 0 or right_delta == 0 or left_delta * right_delta < 0
            is_jump = abs(left_delta) >= threshold or abs(right_delta) >= threshold
            if is_turn and is_jump:
                indices.add(idx)

    filtered = []
    min_gap = max(1, len(values) // 8)
    span = max(values) - min(values) if values else 0
    for idx in sorted(indices):
        if not filtered:
            filtered.append(idx)
            continue
        prev_idx = filtered[-1]
        if idx - prev_idx < min_gap and abs(values[idx] - values[prev_idx]) < max(1.0, span * 0.08):
            continue
        filtered.append(idx)

    if len(filtered) > 7:
        priority = {0, len(values) - 1, max(range(len(values)), key=lambda i: values[i]), min(range(len(values)), key=lambda i: values[i])}
        trimmed = [idx for idx in filtered if idx in priority]
        for idx in filtered:
            if idx not in trimmed and len(trimmed) < 7:
                trimmed.append(idx)
        filtered = sorted(set(trimmed))
    return filtered


def annotate_day_points(ax, times, balances):
    offsets = [18, -24, 24, -18, 30, -30, 16]
    for pos, idx in enumerate(get_day_annotation_indices(balances)):
        y_offset = offsets[pos % len(offsets)]
        ax.annotate(
            format_decimal(balances[idx]),
            (times[idx], balances[idx]),
            textcoords="offset points",
            xytext=(0, y_offset),
            ha='center',
            fontsize=7,
            color="#111827",
            bbox={
                "boxstyle": "round,pad=0.22",
                "facecolor": "#ffffff",
                "edgecolor": "#cbd5e1",
                "linewidth": 0.8
            },
            arrowprops={
                "arrowstyle": "-",
                "color": "#94a3b8",
                "linewidth": 0.8
            },
            clip_on=True
        )


def draw_bot_overview_card(ax, snapshot):
    palette = snapshot["palette"]
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')

    container = FancyBboxPatch(
        (0.04, 0.05),
        0.92,
        0.90,
        boxstyle="round,pad=0.02,rounding_size=0.04",
        linewidth=1.6,
        edgecolor=palette["accent"],
        facecolor=palette["surface"]
    )
    header = FancyBboxPatch(
        (0.04, 0.76),
        0.92,
        0.19,
        boxstyle="round,pad=0.02,rounding_size=0.04",
        linewidth=0,
        facecolor=palette["header"]
    )
    ax.add_patch(container)
    ax.add_patch(header)
    ax.text(0.08, 0.90, snapshot["title"], fontsize=9.8, fontweight='bold', color="#0f172a", va='center')
    ax.text(0.08, 0.81, snapshot["subtitle"], fontsize=7.1, color="#334155", va='center')
    ax.text(
        0.93,
        0.90,
        short_text(snapshot["badge"], 14),
        fontsize=6.8,
        color="#0f172a",
        ha='right',
        va='center',
        bbox={
            "boxstyle": "round,pad=0.25",
            "facecolor": "#ffffff",
            "edgecolor": palette["accent"],
            "linewidth": 0.9
        }
    )
    ax.plot([0.08, 0.92], [0.72, 0.72], color="#e2e8f0", linewidth=1)

    y = 0.66
    for label, value in snapshot["overview_lines"][:6]:
        ax.text(
            0.08,
            y,
            f"{label}: {short_text(value, 27)}",
            fontsize=7.0,
            color="#111827",
            va='top',
            fontfamily='DejaVu Sans Mono'
        )
        y -= 0.10


def get_performance_style(value):
    number = safe_float(value)
    if number is None:
        return {
            "text": "#0f172a",
            "fill": "#ffffff",
            "edge": "#dbe4ee"
        }
    if number >= 0:
        return {
            "text": "#047857",
            "fill": "#ecfdf5",
            "edge": "#6ee7b7"
        }
    return {
        "text": "#b91c1c",
        "fill": "#fef2f2",
        "edge": "#fca5a5"
    }


def draw_metric_box(ax, x, y, width, height, label, value, palette, value_color="#0f172a", scale=1.0):
    shadow = FancyBboxPatch(
        (x + 0.006, y - 0.006),
        width,
        height,
        boxstyle="round,pad=0.02,rounding_size=0.025",
        linewidth=0,
        facecolor="#dbe7f3",
        alpha=0.18
    )
    metric_box = FancyBboxPatch(
        (x, y),
        width,
        height,
        boxstyle="round,pad=0.02,rounding_size=0.025",
        linewidth=1.0,
        edgecolor="#dbe4ee",
        facecolor=palette.get("surface", "#ffffff")
    )
    ax.add_patch(shadow)
    ax.add_patch(metric_box)
    ax.plot(
        [x + 0.03, x + width - 0.03],
        [y + height - 0.04, y + height - 0.04],
        color="#e2e8f0",
        linewidth=0.9
    )
    display_label = short_text(label, 15)
    display_value = short_text(value, 22)
    value_length = len(display_value)
    if value_length > 18:
        value_fontsize = 11.0 * scale
    elif value_length > 14:
        value_fontsize = 11.8 * scale
    else:
        value_fontsize = 12.8 * scale

    label_fontsize = (6.7 if len(display_label) > 12 else 7.0) * scale
    ax.text(x + 0.03, y + height - 0.018, display_label, fontsize=label_fontsize, color="#64748b", va='top')
    ax.text(
        x + 0.03,
        y + 0.028,
        display_value,
        fontsize=value_fontsize,
        color=value_color,
        va='bottom',
        fontweight='bold'
    )


def draw_detail_panel(ax, x, y, width, height, title, items, palette, scale=1.0):
    panel = FancyBboxPatch(
        (x, y),
        width,
        height,
        boxstyle="round,pad=0.02,rounding_size=0.03",
        linewidth=1.0,
        edgecolor="#dbe4ee",
        facecolor="#ffffff"
    )
    ax.add_patch(panel)
    ax.text(x + 0.03, y + height - 0.04, title, fontsize=10.5 * scale, fontweight='bold', color="#0f172a", va='top')
    ax.plot(
        [x + 0.03, x + width - 0.03],
        [y + height - 0.07, y + height - 0.07],
        color=palette["accent"],
        linewidth=1.1,
        alpha=0.22
    )

    if not items:
        ax.text(x + 0.03, y + height - 0.12, "Нет данных", fontsize=8.4 * scale, color="#64748b", va='top')
        return

    is_compact = scale < 0.95
    label_limit = 13 if is_compact else 17
    value_limit = 13 if is_compact else 18
    start_y = y + height - (0.095 if is_compact else 0.10)
    row_gap = min(0.062, max((0.038 if is_compact else 0.032) * scale, (height - 0.115) / max(len(items), 1)))
    current_y = start_y
    for idx, (label, value) in enumerate(items):
        display_label = short_text(label, label_limit)
        display_value = short_text(value, value_limit)
        label_fontsize = ((6.7 if len(display_label) > 11 else 7.0) if is_compact else (7.1 if len(display_label) > 14 else 7.4)) * scale
        value_fontsize = ((7.4 if len(display_value) > 10 else 7.8) if is_compact else (8.4 if len(display_value) > 14 else 8.8)) * scale
        ax.text(x + 0.03, current_y, display_label, fontsize=label_fontsize, color="#64748b", va='center')
        ax.text(
            x + width - 0.03,
            current_y,
            display_value,
            fontsize=value_fontsize,
            color="#0f172a",
            va='center',
            ha='right',
            fontweight='bold'
        )
        if idx < len(items) - 1:
            separator_y = current_y - (row_gap * 0.5)
            ax.plot(
                [x + 0.03, x + width - 0.03],
                [separator_y, separator_y],
                color="#edf2f7",
                linewidth=0.9
            )
        current_y -= row_gap


def draw_highlight_chip(ax, x, y, width, height, label, value, style, scale=1.0):
    chip = FancyBboxPatch(
        (x, y),
        width,
        height,
        boxstyle="round,pad=0.02,rounding_size=0.028",
        linewidth=1.1,
        edgecolor=style["edge"],
        facecolor=style["fill"]
    )
    ax.add_patch(chip)
    display_label = short_text(label, 12)
    display_value = short_text(value, 16)
    value_fontsize = (11.0 if len(display_value) > 12 else 11.8) * scale
    ax.text(
        x + 0.03,
        y + height * 0.50,
        display_label,
        fontsize=7.2 * scale,
        color="#64748b",
        va='center'
    )
    ax.text(
        x + width - 0.03,
        y + height * 0.50,
        display_value,
        fontsize=value_fontsize,
        color=style["text"],
        va='center',
        ha='right',
        fontweight='bold'
    )


def draw_bot_info_card(ax, snapshot, compact=False):
    palette = snapshot["palette"]
    scale = 0.88 if compact else 1.0
    header_title_size = 14.5 if compact else 19.0
    header_subtitle_size = 8.8 if compact else 10.2
    badge_size = 7.8 if compact else 9.0
    meta_note_size = 7.1 if compact else 8.0
    footer_size = 7.1 if compact else 7.8

    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')

    container = FancyBboxPatch(
        (0.04, 0.05),
        0.92,
        0.90,
        boxstyle="round,pad=0.02,rounding_size=0.04",
        linewidth=1.6,
        edgecolor=palette["accent"],
        facecolor="#ffffff"
    )
    header = FancyBboxPatch(
        (0.04, 0.78),
        0.92,
        0.17,
        boxstyle="round,pad=0.02,rounding_size=0.04",
        linewidth=0,
        facecolor=palette["header"]
    )
    ax.add_patch(container)
    ax.add_patch(header)

    ax.text(0.08, 0.89, snapshot["title"], fontsize=header_title_size, fontweight='bold', color="#0f172a", va='center')
    ax.text(0.08, 0.82, snapshot["subtitle"], fontsize=header_subtitle_size, color="#334155", va='center')
    ax.text(
        0.92,
        0.89,
        short_text(snapshot["badge"], 18 if compact else 22),
        fontsize=badge_size,
        color="#0f172a",
        ha='right',
        va='center',
        bbox={
            "boxstyle": "round,pad=0.28",
            "facecolor": "#ffffff",
            "edgecolor": palette["accent"],
            "linewidth": 1.0
        }
    )
    ax.text(
        0.92,
        0.82,
        snapshot.get("source_note", "Текущий снимок Bybit"),
        fontsize=meta_note_size,
        color="#64748b",
        ha='right',
        va='center'
    )

    pnl_style = get_performance_style(snapshot.get("pnl_usdt"))
    pnl_percent_style = get_performance_style(snapshot.get("pnl_percent_value"))
    draw_highlight_chip(
        ax,
        0.08,
        0.665,
        0.36,
        0.075,
        snapshot.get("highlight_primary_label", "Доход"),
        snapshot.get("highlight_primary_value") or format_usdt(snapshot.get("pnl_usdt"), fallback="N/A"),
        pnl_style,
        scale=scale
    )
    draw_highlight_chip(
        ax,
        0.52,
        0.665,
        0.36,
        0.075,
        snapshot.get("highlight_secondary_label", "Доход %"),
        snapshot.get("highlight_secondary_value") or format_percent(snapshot.get("pnl_percent_value"), scale=1, fallback="N/A"),
        pnl_percent_style,
        scale=scale
    )

    metric_items = list(snapshot["metrics"][:4])
    while len(metric_items) < 4:
        metric_items.append(("Метрика", "N/A"))
    metric_positions = [
        (0.08, 0.515), (0.52, 0.515),
        (0.08, 0.365), (0.52, 0.365)
    ]
    for (label, value), (x, y) in zip(metric_items, metric_positions):
        metric_color = "#0f172a"
        if label in {"P&L", "Общий P&L", "Реал. P&L", "Нереал. P&L", "% P&L", "% APR"}:
            source_value = snapshot.get("pnl_percent_value") if "%" in label else snapshot.get("pnl_usdt")
            metric_color = get_performance_style(source_value)["text"]
        draw_metric_box(ax, x, y, 0.36, 0.11, label, value, palette, value_color=metric_color, scale=scale)

    details = list(snapshot.get("details") or [])
    detail_limit = 6 if compact else 8
    details = details[:detail_limit]
    column_size = (len(details) + 1) // 2
    left_column = details[:column_size]
    right_column = details[column_size:]
    left_title, right_title = snapshot.get("detail_titles", ("Параметры", "Детали"))
    detail_y = 0.115 if compact else 0.105
    detail_height = 0.205 if compact else 0.22
    draw_detail_panel(ax, 0.08, detail_y, 0.36, detail_height, left_title, left_column, palette, scale=scale)
    draw_detail_panel(ax, 0.52, detail_y, 0.36, detail_height, right_title, right_column, palette, scale=scale)

    footer_left_text = snapshot.get("footer_left", "Источник: Bybit")
    if compact:
        footer_left_text = short_text(footer_left_text, 32)
    ax.text(0.08, 0.08, footer_left_text, fontsize=footer_size, color="#64748b", va='center')
    ax.text(
        0.92,
        0.08,
        f"Обновлено: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
        fontsize=footer_size,
        color="#64748b",
        ha='right',
        va='center'
    )


def draw_bot_day_chart(ax, selected_date, snapshot, history_points):
    ax.set_facecolor("#ffffff")
    for spine in ax.spines.values():
        spine.set_color("#cbd5e1")
        spine.set_linewidth(1.0)

    ax.grid(True, which='both', linestyle='--', linewidth=0.6, alpha=0.35, color="#94a3b8")
    ax.set_title("Прибыль и баланс за день", fontsize=11.2, fontweight='bold', loc='left', pad=12)
    ax.text(
        1.0,
        1.02,
        selected_date.strftime('%d.%m.%Y'),
        transform=ax.transAxes,
        fontsize=8.3,
        color="#64748b",
        va='bottom',
        ha='right'
    )

    if not history_points:
        ax.set_xticks([])
        ax.set_yticks([])
        ax.text(
            0.5,
            0.58,
            "История бота\nещё не накоплена",
            ha='center',
            va='center',
            transform=ax.transAxes,
            fontsize=11.0,
            fontweight='bold',
            color="#0f172a"
        )
        ax.text(
            0.5,
            0.40,
            "После нескольких обновлений БД\nздесь появится дневной график.",
            ha='center',
            va='center',
            transform=ax.transAxes,
            fontsize=8.4,
            color="#64748b"
        )
        return

    equity_points = [(item["time"], item["equity_usdt"]) for item in history_points if item["equity_usdt"] is not None]
    pnl_points = [(item["time"], item["pnl_usdt"]) for item in history_points if item["pnl_usdt"] is not None]

    if not equity_points and not pnl_points:
        ax.set_xticks([])
        ax.set_yticks([])
        ax.text(
            0.5,
            0.5,
            "Для этого бота пока нет\nчисловых значений баланса и прибыли.",
            ha='center',
            va='center',
            transform=ax.transAxes,
            fontsize=9.2,
            color="#64748b"
        )
        return

    lines = []
    labels = []
    if equity_points:
        equity_times = [item[0] for item in equity_points]
        equity_values = [item[1] for item in equity_points]
        line_kwargs = {
            "linewidth": 2.2,
            "color": snapshot["palette"]["accent"],
            "label": "Баланс бота",
            "solid_capstyle": "round",
            "solid_joinstyle": "round"
        }
        if len(equity_points) == 1:
            line_kwargs.update({"marker": "o", "markersize": 4})
            ax.set_xlim(equity_times[0] - timedelta(minutes=30), equity_times[0] + timedelta(minutes=30))
        line_balance, = ax.plot(equity_times, equity_values, **line_kwargs)
        ax.annotate(
            format_decimal(equity_values[-1]),
            (equity_times[-1], equity_values[-1]),
            textcoords="offset points",
            xytext=(0, 12),
            ha='center',
            fontsize=7,
            bbox={
                "boxstyle": "round,pad=0.22",
                "facecolor": "#ffffff",
                "edgecolor": "#cbd5e1",
                "linewidth": 0.9
            }
        )
        lines.append(line_balance)
        labels.append("Баланс бота")
        ax.set_ylabel('Баланс (USDT)', fontsize=8.5, color="#334155")

    pnl_axis = None
    if pnl_points:
        pnl_axis = ax.twinx()
        pnl_axis.grid(False)
        pnl_axis.patch.set_alpha(0)
        pnl_axis.spines['right'].set_color("#cbd5e1")
        pnl_axis.spines['right'].set_linewidth(1.0)
        pnl_axis.spines['top'].set_visible(False)
        pnl_axis.spines['left'].set_visible(False)
        pnl_axis.spines['bottom'].set_visible(False)

        pnl_times = [item[0] for item in pnl_points]
        pnl_values = [item[1] for item in pnl_points]
        pnl_color = "#059669" if pnl_values[-1] >= 0 else "#b45309"
        pnl_kwargs = {
            "linewidth": 2.0,
            "color": pnl_color,
            "label": "P&L",
            "solid_capstyle": "round",
            "solid_joinstyle": "round"
        }
        if len(pnl_points) == 1:
            pnl_kwargs.update({"marker": "o", "markersize": 3.5})
            if not equity_points:
                ax.set_xlim(pnl_times[0] - timedelta(minutes=30), pnl_times[0] + timedelta(minutes=30))
        line_pnl, = pnl_axis.plot(pnl_times, pnl_values, **pnl_kwargs)
        pnl_axis.annotate(
            format_decimal(pnl_values[-1]),
            (pnl_times[-1], pnl_values[-1]),
            textcoords="offset points",
            xytext=(0, -16),
            ha='center',
            fontsize=7,
            bbox={
                "boxstyle": "round,pad=0.22",
                "facecolor": "#ffffff",
                "edgecolor": "#cbd5e1",
                "linewidth": 0.9
            }
        )
        lines.append(line_pnl)
        labels.append("P&L")
        pnl_axis.set_ylabel('P&L (USDT)', fontsize=8.5, color="#475569")
        pnl_axis.tick_params(axis='y', labelsize=8, colors="#475569")

    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    ax.tick_params(axis='y', labelsize=8, colors="#334155")
    ax.margins(x=0.04, y=0.18)

    if lines:
        ax.legend(lines, labels, loc='upper left', fontsize=7.4, frameon=False)

    summary_lines = []
    if snapshot.get("investment_usdt") is not None:
        summary_lines.append(f"База: {format_usdt(snapshot['investment_usdt'])}")
    if equity_points:
        summary_lines.append(f"Баланс: {format_usdt(equity_points[-1][1])}")
    if pnl_points:
        summary_lines.append(f"P&L: {format_usdt(pnl_points[-1][1])}")

    if summary_lines:
        ax.text(
            0.98,
            0.04,
            "\n".join(summary_lines[:3]),
            transform=ax.transAxes,
            ha='right',
            va='bottom',
            fontsize=7.4,
            color="#0f172a",
            bbox={
                "boxstyle": "round,pad=0.28",
                "facecolor": "#ffffff",
                "edgecolor": "#dbe4ee",
                "linewidth": 0.9
            }
        )


def build_bot_card_caption(snapshot):
    lines = [
        snapshot["title"],
        f"Тип: {snapshot['subtitle']}",
        f"Режим: {snapshot['badge']}"
    ]

    seen_labels = set()
    for label, value in snapshot["metrics"]:
        if label in seen_labels:
            continue
        seen_labels.add(label)
        lines.append(f"{label}: {value}")

    for label, value in snapshot["details"]:
        if label in seen_labels or label in {"Режим"}:
            continue
        seen_labels.add(label)
        lines.append(f"{label}: {value}")
        if len(lines) >= 10:
            break

    return "\n".join(lines[:10])


def get_top_bot_profit_value(record):
    is_active = bool(safe_int(record.get("is_active")))
    if is_active:
        return safe_float(record.get("pnl_usdt"))
    for key in ("final_profit_usdt", "pnl_usdt"):
        value = safe_float(record.get(key))
        if value is not None:
            return value
    investment_value = safe_float(record.get("investment_usdt"))
    settlement_value = safe_float(record.get("settlement_assets_usdt"))
    if investment_value is not None and settlement_value is not None:
        return settlement_value - investment_value
    return None


def normalize_top_sort_mode(sort_mode):
    if sort_mode in TOP_SORT_MODES:
        return sort_mode
    return "earnings"


def get_top_bots_image_path(sort_mode, page=0):
    normalized_mode = normalize_top_sort_mode(sort_mode)
    page_index = max(0, safe_int(page) or 0)
    if normalized_mode == "earnings" and page_index == 0:
        return TOP_BOTS_IMAGE_FILE
    return os.path.join(CACHE_DIR, f"top_bots_{normalized_mode}_p{page_index + 1}.png")


def get_top_bot_percent_value(record):
    investment_value = safe_float(record.get("investment_usdt"))
    profit_value = get_top_bot_profit_value(record)
    calculated_value = calculate_profit_percent(profit_value, investment_value)
    if calculated_value is not None:
        return calculated_value
    return safe_float(record.get("pnl_percent"))


def get_top_bot_pnl_value(record):
    pnl_value = safe_float(record.get("pnl_usdt"))
    if pnl_value is not None:
        return pnl_value
    return get_top_bot_profit_value(record)


def get_top_bot_metric_value(record, sort_mode):
    normalized_mode = normalize_top_sort_mode(sort_mode)
    if normalized_mode == "percent":
        return get_top_bot_percent_value(record)
    if normalized_mode == "pnl":
        return get_top_bot_pnl_value(record)
    return get_top_bot_profit_value(record)


def format_top_bot_metric(sort_mode, value, fallback="N/A"):
    normalized_mode = normalize_top_sort_mode(sort_mode)
    if normalized_mode == "percent":
        return format_percent(value, scale=1, fallback=fallback)
    return format_usdt(value, fallback=fallback)


def get_top_bot_rows(limit=20, sort_mode="earnings"):
    if not USE_DB:
        return []
    ensure_db_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM bot_archive")
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    normalized_mode = normalize_top_sort_mode(sort_mode)
    sorted_rows = sorted(
        rows,
        key=lambda record: (
            1 if safe_int(record.get("is_active")) else 0,
            get_top_bot_metric_value(record, normalized_mode) is None,
            -(get_top_bot_metric_value(record, normalized_mode) or 0.0),
            -(normalize_epoch_timestamp(record.get("ended_ts"))
              or normalize_epoch_timestamp(record.get("created_ts"))
              or 0)
        )
    )
    return sorted_rows[:int(limit)]


def paginate_top_rows(rows, page=0, page_size=TOP_BOTS_PAGE_SIZE):
    total_rows = len(rows or [])
    page_size_value = max(1, safe_int(page_size) or TOP_BOTS_PAGE_SIZE)
    total_pages = max(1, math.ceil(total_rows / page_size_value)) if total_rows else 1
    page_index = min(max(0, safe_int(page) or 0), total_pages - 1)
    start_index = page_index * page_size_value
    end_index = start_index + page_size_value
    return (rows or [])[start_index:end_index], page_index, total_pages, start_index


def build_top_bot_snapshot(record, index, sort_mode="earnings"):
    raw_bot_data = None
    raw_json = record.get("raw_json")
    if raw_json:
        try:
            parsed = json.loads(raw_json)
            if isinstance(parsed, dict):
                raw_bot_data = parsed
        except Exception:
            raw_bot_data = None

    snapshot = build_bot_snapshot(raw_bot_data or {}, index)
    normalized_mode = normalize_top_sort_mode(sort_mode)
    mode_meta = TOP_SORT_MODES[normalized_mode]
    bot_type = record.get("bot_type") or snapshot.get("bot_type") or "UNKNOWN"
    symbol = record.get("symbol") or snapshot.get("symbol") or record.get("title") or f"Бот {index + 1}"
    profit_value = get_top_bot_profit_value(record)
    pnl_value = get_top_bot_pnl_value(record)
    metric_value = get_top_bot_metric_value(record, normalized_mode)
    investment_value = safe_float(record.get("investment_usdt"))
    if investment_value is None:
        investment_value = snapshot.get("investment_usdt")
    pnl_percent_value = get_top_bot_percent_value(record)

    is_active = bool(safe_int(record.get("is_active")))
    status_value = "Активен" if is_active else "Закрыт"
    time_label = "Время запуска" if is_active else "Закрыт"
    time_value = (
        snapshot.get("started_at_text") or "N/A"
        if is_active else
        format_short_datetime(record.get("ended_ts"))
    )
    reason_value = (
        record.get("close_code")
        or record.get("close_reason")
        or record.get("display_status")
        or record.get("status")
        or "N/A"
    )

    snapshot["bot_type"] = bot_type
    snapshot["symbol"] = symbol
    snapshot["title"] = f"#{index + 1} {symbol}"
    snapshot["subtitle"] = get_bot_type_name(bot_type)
    snapshot["badge"] = record.get("badge") or snapshot.get("badge") or get_bot_button_kind(bot_type)
    snapshot["investment_usdt"] = investment_value
    snapshot["pnl_usdt"] = profit_value if profit_value is not None else pnl_value
    snapshot["pnl_percent_value"] = pnl_percent_value
    if investment_value is not None and profit_value is not None:
        snapshot["equity_usdt"] = investment_value + profit_value
    snapshot["palette"] = get_card_palette(bot_type, profit_value if profit_value is not None else pnl_value)
    snapshot["source_note"] = "Архив и история Bybit" if not is_active else "Активный бот Bybit"
    snapshot["footer_left"] = (
        f"Статус: {status_value}"
        if is_active else
        f"Причина: {short_text(reason_value, 28)}"
    )
    snapshot["highlight_primary_label"] = mode_meta["overview_label"]
    snapshot["highlight_primary_value"] = format_top_bot_metric(normalized_mode, metric_value, fallback="N/A")
    if normalized_mode == "percent":
        snapshot["highlight_secondary_label"] = "Итог"
        snapshot["highlight_secondary_value"] = format_usdt(profit_value, fallback="N/A")
    else:
        snapshot["highlight_secondary_label"] = "% ROI"
        snapshot["highlight_secondary_value"] = format_percent(pnl_percent_value, scale=1, fallback="N/A")
    top_details = [
        ("Статус", status_value),
        (time_label, time_value),
        ("Причина", reason_value)
    ]
    merged_details = []
    for label, value in top_details + list(snapshot.get("details") or []):
        if label in {item[0] for item in merged_details}:
            continue
        merged_details.append((label, value))
    snapshot["details"] = merged_details[:10]
    overview_lines = [
        ("Статус", status_value),
        (mode_meta["overview_label"], format_top_bot_metric(normalized_mode, metric_value, fallback="N/A")),
        ("Итог", format_usdt(profit_value, fallback="N/A")),
        ("P&L", format_usdt(pnl_value, fallback="N/A")),
        ("% ROI", format_percent(pnl_percent_value, scale=1, fallback="N/A")),
        ("Вложено", format_usdt(investment_value, fallback="N/A"))
    ]
    snapshot["overview_lines"] = overview_lines[:6]
    return snapshot


def build_top_bots_markup(active_mode="earnings", page=0, total_pages=1):
    normalized_mode = normalize_top_sort_mode(active_mode)
    markup = types.InlineKeyboardMarkup(row_width=3)
    buttons = []
    for mode_key, mode_meta in TOP_SORT_MODES.items():
        label = mode_meta["button"]
        if mode_key == normalized_mode:
            label = f"• {label}"
        buttons.append(types.InlineKeyboardButton(label, callback_data=f"top_view_{mode_key}_0"))
    markup.add(*buttons)
    if total_pages > 1:
        nav_buttons = []
        previous_page = max(0, page - 1)
        next_page = min(total_pages - 1, page + 1)
        nav_buttons.append(types.InlineKeyboardButton("←", callback_data=f"top_view_{normalized_mode}_{previous_page}"))
        nav_buttons.append(types.InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="top_noop"))
        nav_buttons.append(types.InlineKeyboardButton("→", callback_data=f"top_view_{normalized_mode}_{next_page}"))
        markup.add(*nav_buttons)
    return markup


def generate_top_bots_image(sort_mode="earnings", force_refresh=False, page=0):
    normalized_mode = normalize_top_sort_mode(sort_mode)
    all_rows = get_top_bot_rows(limit=20, sort_mode=normalized_mode)
    if not all_rows:
        return None, "В архиве ботов пока нет данных.", [], 0, 1

    page_rows, page_index, total_pages, start_index = paginate_top_rows(all_rows, page=page, page_size=TOP_BOTS_PAGE_SIZE)

    image_path = get_top_bots_image_path(normalized_mode, page=page_index)
    if os.path.exists(image_path) and not force_refresh:
        return image_path, None, page_rows, page_index, total_pages

    os.makedirs(CACHE_DIR, exist_ok=True)
    grid_rows = max(1, math.ceil(len(page_rows) / 2))
    fig_height = 4.6 * grid_rows + 0.9
    fig = plt.figure(figsize=(13.2, fig_height))
    fig.patch.set_facecolor("#eef2f7")
    fig.suptitle(
        (
            f"Топ 20 ботов по {TOP_SORT_MODES[normalized_mode]['title']} "
            f"• стр. {page_index + 1}/{total_pages} • {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ),
        fontsize=16,
        fontweight='bold',
        color="#0f172a",
        y=0.992
    )
    gs = fig.add_gridspec(
        grid_rows,
        2,
        left=0.04,
        right=0.97,
        top=0.975,
        bottom=0.02,
        wspace=0.09,
        hspace=0.10
    )

    axes = [fig.add_subplot(gs[row, col]) for row in range(grid_rows) for col in range(2)]
    for index, ax in enumerate(axes):
        ax.axis('off')
        if index >= len(page_rows):
            ax.set_visible(False)
            continue
        snapshot = build_top_bot_snapshot(page_rows[index], start_index + index, sort_mode=normalized_mode)
        draw_bot_info_card(ax, snapshot, compact=True)

    plt.savefig(image_path, dpi=200)
    plt.close(fig)
    return image_path, None, page_rows, page_index, total_pages


def build_top_bots_caption(top_rows, sort_mode="earnings", page=0, total_pages=1):
    normalized_mode = normalize_top_sort_mode(sort_mode)
    mode_meta = TOP_SORT_MODES[normalized_mode]
    lines = [
        f"Топ 20 ботов по {mode_meta['title']}",
        f"Страница: {page + 1}/{total_pages} • Обновлено: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ]
    for index, row in enumerate(top_rows[:5], start=1):
        symbol = row.get("symbol") or row.get("title") or f"Бот {index}"
        metric_value = get_top_bot_metric_value(row, normalized_mode)
        lines.append(
            f"{index}. {symbol}: {mode_meta['caption_label']} "
            f"{format_top_bot_metric(normalized_mode, metric_value, fallback='N/A')}"
        )
    return "\n".join(lines[:7])


def show_top_bots(chat_id, sort_mode="earnings", page=0, message_id=None, refresh_archive=False):
    normalized_mode = normalize_top_sort_mode(sort_mode)
    if refresh_archive and USE_DB:
        sync_bot_archive(force=True, include_active=True, include_history=True)
    file_path, error, top_rows, page_index, total_pages = generate_top_bots_image(
        sort_mode=normalized_mode,
        force_refresh=True,
        page=page
    )
    if error:
        if message_id is None:
            bot.send_message(chat_id, error)
        else:
            bot.edit_message_caption(chat_id=chat_id, message_id=message_id, caption=error)
        return False
    caption = build_top_bots_caption(top_rows, sort_mode=normalized_mode, page=page_index, total_pages=total_pages)
    markup = build_top_bots_markup(active_mode=normalized_mode, page=page_index, total_pages=total_pages)
    if message_id is None:
        bot.send_photo(chat_id, types.InputFile(file_path), caption=caption, reply_markup=markup)
    else:
        bot.edit_message_media(types.InputMediaPhoto(types.InputFile(file_path)), chat_id=chat_id, message_id=message_id)
        bot.edit_message_caption(chat_id=chat_id, message_id=message_id, caption=caption, reply_markup=markup)
    return True


def generate_bot_card_image(selected_date, bot_index, force_refresh=False, bots_data=None):
    if bots_data is None:
        bots_data = fetch_bot_list_data()[:6]
    if bot_index < 0 or bot_index >= len(bots_data):
        return None, "Карточка бота недоступна."

    snapshot = build_bot_snapshot(bots_data[bot_index], bot_index)
    card_path = get_bot_card_path(selected_date, bot_index)
    if os.path.exists(card_path) and not force_refresh:
        return card_path, None

    history_points = get_bot_day_history(selected_date, snapshot)
    if (
        not history_points
        and selected_date == datetime.now().date()
        and (snapshot["equity_usdt"] is not None or snapshot["pnl_usdt"] is not None)
    ):
        history_points = [{
            "time": datetime.now().replace(second=0, microsecond=0),
            "investment_usdt": snapshot["investment_usdt"],
            "pnl_usdt": snapshot["pnl_usdt"],
            "equity_usdt": snapshot["equity_usdt"]
        }]

    fig = plt.figure(figsize=(12.6, 6.0))
    fig.patch.set_facecolor("#edf2f7")
    ax = fig.add_axes([0.03, 0.05, 0.50, 0.90])
    ax_chart = fig.add_axes([0.56, 0.10, 0.41, 0.80])
    draw_bot_info_card(ax, snapshot, compact=False)
    draw_bot_day_chart(ax_chart, selected_date, snapshot, history_points)
    plt.savefig(card_path, dpi=220)
    plt.close(fig)
    return card_path, None


def build_graph_message_markup(selected_date, bots_data=None, viewer_user_id=None):
    if bots_data is None:
        bots_data = fetch_bot_list_data()[:6]
    markup, _ = generate_calendar_markup(selected_date.year, selected_date.month)
    if markup is None:
        markup = types.InlineKeyboardMarkup(row_width=2)

    bot_buttons = []
    date_token = date_to_token(selected_date)
    for idx, bot_data in enumerate(bots_data[:6]):
        snapshot = build_bot_snapshot(bot_data, idx)
        bot_buttons.append(
            types.InlineKeyboardButton(
                snapshot["button_label"],
                callback_data=f"graph_bot_{date_token}_{idx}"
            )
        )

    for index in range(0, len(bot_buttons), 2):
        markup.add(*bot_buttons[index:index + 2])

    if viewer_user_id is not None and is_admin(viewer_user_id):
        markup.add(types.InlineKeyboardButton("Назад в админку", callback_data="graph_admin_back"))
    return markup


def build_bot_card_markup(selected_date, active_index, bots_data=None, viewer_user_id=None):
    if bots_data is None:
        bots_data = fetch_bot_list_data()[:6]
    markup = types.InlineKeyboardMarkup(row_width=2)
    date_token = date_to_token(selected_date)
    bot_buttons = []
    for idx, bot_data in enumerate(bots_data[:6]):
        snapshot = build_bot_snapshot(bot_data, idx)
        label = snapshot["button_label"]
        if idx == active_index:
            label = f"• {short_text(label, 22)}"
        bot_buttons.append(
            types.InlineKeyboardButton(
                label,
                callback_data=f"graph_bot_{date_token}_{idx}"
            )
        )

    for index in range(0, len(bot_buttons), 2):
        markup.add(*bot_buttons[index:index + 2])

    back_buttons = [types.InlineKeyboardButton("Назад к графику", callback_data=f"graph_overview_{date_token}")]
    if viewer_user_id is not None and is_admin(viewer_user_id):
        back_buttons.append(types.InlineKeyboardButton("Назад в админку", callback_data="graph_admin_back"))
    markup.add(*back_buttons)
    return markup


def generate_graph_for_date(selected_date=None, bot_obj=None, force_refresh=False, bots_data=None):
    if bots_data is None:
        bots_data = fetch_bot_list_data()[:6]

    rows = get_effective_balance_history()
    if not rows:
        return None, "Нет данных."
    if selected_date is None:
        all_dates = sorted(list(set(r[0].date() for r in rows)))
        if not all_dates:
            return None, "Нет данных."
        selected_date = all_dates[-1]
    graph_filename = get_overview_graph_path(selected_date)
    if os.path.exists(graph_filename) and not force_refresh:
        return graph_filename, None

    ref_date = selected_date
    day_rows = [r for r in rows if r[0].date() == selected_date and r[1] is not None]
    if not day_rows:
        all_dates = sorted(list(set(r[0].date() for r in rows)))
        selected_date = all_dates[-1]
        day_rows = [r for r in rows if r[0].date() == selected_date and r[1] is not None]
    day_rows.sort(key=lambda x: x[0])
    times = [r[0] for r in day_rows]
    balances_usdt = [r[1] for r in day_rows]

    # Группировка по дням для 30-дневного графика (от ref_date - 29 дней до ref_date)
    daily_balances = {}
    for r in rows:
        d = r[0].date()
        if d <= ref_date and r[1] is not None:
            daily_balances.setdefault(d, []).append(r[1])
    all_dates_sorted = sorted([d for d in daily_balances.keys() if d <= ref_date])
    last_30_days = [d for d in all_dates_sorted if d >= (ref_date - timedelta(days=29))]

    avg_30, max_30, min_30, dates_30 = [], [], [], []
    for d in last_30_days:
        vals = daily_balances[d]
        vals = [v for v in vals if v is not None]
        if not vals:
            continue
        avg_30.append(sum(vals) / len(vals))
        max_30.append(max(vals))
        min_30.append(min(vals))
        dates_30.append(d)

    # Группировка по месяцам для годового графика (от ref_date - 364 дней до ref_date)
    monthly_balances = {}
    for d, vals in daily_balances.items():
        if d <= ref_date and d >= (ref_date - timedelta(days=364)):
            m = d.replace(day=1)
            monthly_balances.setdefault(m, []).extend(vals)
    avg_year, max_year, min_year, dates_year = [], [], [], []
    if monthly_balances:
        for m in sorted(monthly_balances.keys()):
            vs = [v for v in monthly_balances[m] if v is not None]
            if not vs:
                continue
            avg_year.append(sum(vs) / len(vs))
            max_year.append(max(vs))
            min_year.append(min(vs))
            dates_year.append(m)
    fig = plt.figure(figsize=(17.2, 10.4))
    fig.patch.set_facecolor("#eef2f7")
    gs = fig.add_gridspec(
        3,
        3,
        width_ratios=[4.5, 1.35, 1.35],
        height_ratios=[1, 1, 1],
        left=0.04,
        right=0.98,
        top=0.965,
        bottom=0.055,
        wspace=0.12,
        hspace=0.24
    )
    ax_day = fig.add_subplot(gs[0, 0])
    ax_30 = fig.add_subplot(gs[1, 0])
    ax_year = fig.add_subplot(gs[2, 0])
    ax_bot = [fig.add_subplot(gs[i, j]) for i, j in [(0, 1), (0, 2), (1, 1), (1, 2), (2, 1), (2, 2)]]
    for a in ax_bot:
        a.axis('off')
        a.set_facecolor("#eef2f7")
    y_locator = MaxNLocator(nbins=5)
    for ax in [ax_day, ax_30, ax_year]:
        ax.set_facecolor("#ffffff")
        ax.yaxis.set_major_locator(y_locator)
        ax.grid(True, which='both', linestyle='--', linewidth=0.6, alpha=0.45, color="#94a3b8")
        for spine in ax.spines.values():
            spine.set_color("#cbd5e1")
            spine.set_linewidth(1.0)

    ax_day.plot(
        times,
        balances_usdt,
        linestyle='-',
        linewidth=2.3,
        color="#dc2626",
        label='Баланс (текущий день)',
        antialiased=True,
        solid_capstyle='round',
        solid_joinstyle='round'
    )
    annotate_day_points(ax_day, times, balances_usdt)
    ax_day.margins(x=0.02, y=0.14)
    ax_day.set_ylabel('Баланс (USDT)', fontsize=9)
    ax_day.set_title(f'Баланс за {selected_date.strftime("%d_%m_%Y")}', fontsize=11, fontweight='bold')
    ax_day.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
    ax_day.tick_params(axis='x', rotation=45, labelsize=8)
    ax_day.tick_params(axis='y', labelsize=8)
    ax_day.legend(fontsize=7)
    draw_summary_box(ax_day, "День", balances_usdt, anchor=(0.985, 0.18))

    if dates_30:
        ax_30.plot(
            dates_30,
            avg_30,
            linestyle='-',
            linewidth=2.0,
            color="#2563eb",
            label='Средний баланс (30 дней)',
            antialiased=True,
            solid_capstyle='round',
            solid_joinstyle='round'
        )
        ax_30.annotate(
            format_decimal(avg_30[-1]),
            (dates_30[-1], avg_30[-1]),
            textcoords="offset points",
            xytext=(0, 14),
            ha='center',
            fontsize=7,
            bbox={
                "boxstyle": "round,pad=0.25",
                "facecolor": "#ffffff",
                "edgecolor": "#bfdbfe",
                "linewidth": 0.9
            }
        )
        draw_summary_box(ax_30, "30 дней", avg_30)
        ax_30.set_ylabel('Средний баланс (USDT)', fontsize=9)
        ax_30.set_title('Средний баланс за последние 30 дней', fontsize=11, fontweight='bold')
        ax_30.xaxis.set_major_formatter(mdates.DateFormatter('%d\n%b'))
        ax_30.tick_params(axis='x', rotation=45, labelsize=8)
        ax_30.tick_params(axis='y', labelsize=8)
        ax_30.legend(fontsize=7)
    else:
        ax_30.text(0.5, 0.5, 'Нет данных за последние 30 дней', ha='center', va='center', transform=ax_30.transAxes,
                   fontsize=9)
    if dates_year:
        ax_year.plot(
            dates_year,
            avg_year,
            linestyle='-',
            linewidth=2.0,
            color="#059669",
            label='Средний баланс (год)',
            antialiased=True,
            solid_capstyle='round',
            solid_joinstyle='round'
        )
        ax_year.annotate(
            format_decimal(avg_year[-1]),
            (dates_year[-1], avg_year[-1]),
            textcoords="offset points",
            xytext=(0, 14),
            ha='center',
            fontsize=7,
            bbox={
                "boxstyle": "round,pad=0.25",
                "facecolor": "#ffffff",
                "edgecolor": "#a7f3d0",
                "linewidth": 0.9
            }
        )
        draw_summary_box(ax_year, "Год", avg_year)
        ax_year.set_ylabel('Средний баланс (USDT)', fontsize=9)
        ax_year.set_title('Средний баланс за последний год', fontsize=11, fontweight='bold')
        ax_year.xaxis.set_major_formatter(mdates.DateFormatter('%b\n%Y'))
        ax_year.tick_params(axis='x', rotation=45, labelsize=8)
        ax_year.tick_params(axis='y', labelsize=8)
        ax_year.legend(fontsize=7)
    else:
        ax_year.text(0.5, 0.5, 'Нет данных за последний год', ha='center', va='center', transform=ax_year.transAxes,
                     fontsize=9)
    for i, axx in enumerate(ax_bot):
        if i >= len(bots_data):
            axx.set_visible(False)
            continue
        bot_snapshot = build_bot_snapshot(bots_data[i], i)
        draw_bot_overview_card(axx, bot_snapshot)
    plt.savefig(graph_filename, dpi=300)
    plt.close()
    return graph_filename, None


def generate_all_graphs():
    dates = get_all_dates()
    count_new = 0
    for d in dates:
        graph_file = get_overview_graph_path(d)
        if not os.path.exists(graph_file):
            _, err = generate_graph_for_date(d)
            if err is None:
                count_new += 1
    return count_new


def wait_until_next_interval(minutes, run_token=None):
    interval = max(1, int(minutes))
    now = datetime.now()
    current_slot = datetime.strptime(get_interval_slot_key(interval, dt=now), '%Y-%m-%d %H:%M:%S')
    target = current_slot + timedelta(minutes=interval)
    while True:
        if stop_threads:
            return False
        if run_token is not None and run_token != thread_run_token:
            return False
        delta = (target - datetime.now()).total_seconds()
        if delta <= 0:
            return True
        sleep(min(delta, 1.0))


# ------------------ ЦИКЛЫ ОБНОВЛЕНИЯ ------------------

threads_started = False


def start_threads():
    global db_update_thread, balance_send_thread, market_alert_thread
    global stop_threads, threads_started, thread_run_token
    if threads_started:
        return
    stop_threads = False
    thread_run_token += 1
    run_token = thread_run_token
    db_update_thread = threading.Thread(target=db_update_loop, args=(run_token,), daemon=True)
    balance_send_thread = threading.Thread(target=balance_send_loop, args=(run_token,), daemon=True)
    market_alert_thread = threading.Thread(target=market_alert_loop, args=(run_token,), daemon=True)
    db_update_thread.start()
    balance_send_thread.start()
    market_alert_thread.start()
    threads_started = True


def stop_all_threads():
    global stop_threads, threads_started, thread_run_token
    stop_threads = True
    thread_run_token += 1
    threads_started = False


def db_update_loop(run_token):
    while not stop_threads and run_token == thread_run_token:
        try:
            wait_minutes = min(max(1, int(db_update_interval)), BOT_ARCHIVE_SYNC_INTERVAL_MINUTES)
            if claim_schedule_slot("db_slot", db_update_interval):
                fetch_balance()
            sync_bot_archive(include_active=False, include_history=True)
            repair_bot_archive_metrics()
            dispatch_active_bot_risk_alerts()
            dispatch_bot_close_notifications()
        except Exception:
            logging.exception("Ошибка цикла обновления БД")
        if not wait_until_next_interval(wait_minutes, run_token=run_token):
            break


def balance_send_loop(run_token):
    while not stop_threads and run_token == thread_run_token:
        try:
            if claim_schedule_slot("balance_slot", balance_send_interval):
                balance_info = fetch_balance(add_to_db=False)
                if isinstance(balance_info, str) and chat_id:
                    try:
                        bot.send_message(chat_id, balance_info)
                    except Exception:
                        pass
        except Exception:
            logging.exception("Ошибка цикла отправки баланса")
        if not wait_until_next_interval(balance_send_interval, run_token=run_token):
            break


# ------------------ АДМИН-ПАНЕЛЬ ------------------

def is_admin(user_id):
    return user_id in admins


@bot.message_handler(commands=['admin'])
@handler_guard
def admin_panel(message):
    if message.chat.type != 'private':
        return
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, MESSAGES['admin_no_access'])
        return
    bot.send_message(message.chat.id, MESSAGES['admin_panel_title'], reply_markup=get_admin_panel())


def get_admin_panel():
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Изменить TOKEN", callback_data="change_token"),
        types.InlineKeyboardButton("Изменить cookies", callback_data="change_cookies")
    )
    markup.add(
        types.InlineKeyboardButton("Скачать базу данных", callback_data="download_db"),
        types.InlineKeyboardButton("Показать настройки", callback_data="show_config")
    )
    markup.add(
        types.InlineKeyboardButton("Интервал БД", callback_data="change_db_interval"),
        types.InlineKeyboardButton("Интервал баланса", callback_data="change_balance_interval")
    )
    markup.add(
        types.InlineKeyboardButton("Добавить админа", callback_data="add_admin"),
        types.InlineKeyboardButton("Удалить админа", callback_data="remove_admin")
    )
    markup.add(
        types.InlineKeyboardButton("Перечитать конфиг", callback_data="reload_bot")
    )
    markup.add(
        types.InlineKeyboardButton("Уведомления", callback_data="notification_settings")
    )
    markup.add(
        types.InlineKeyboardButton("Отчёт день", callback_data="report_day"),
        types.InlineKeyboardButton("Отчёт неделя", callback_data="report_week")
    )
    markup.add(
        types.InlineKeyboardButton("Миграция Excel -> DB", callback_data="migrate_excel_to_db")
    )
    markup.add(
        types.InlineKeyboardButton("Генерация графиков", callback_data="generate_all_graphs")
    )
    return markup


def format_notification_state(value):
    return "ON" if value else "OFF"


def build_notification_settings_text():
    settings = get_notification_settings()
    lines = ["Настройки уведомлений", ""]
    for key, label in NOTIFICATION_LABELS.items():
        lines.append(f"{label}: {format_notification_state(settings.get(key, False))}")
    return "\n".join(lines)


def get_notification_settings_panel():
    settings = get_notification_settings()
    markup = types.InlineKeyboardMarkup(row_width=2)
    buttons = []
    for key, label in NOTIFICATION_LABELS.items():
        state = format_notification_state(settings.get(key, False))
        buttons.append(
            types.InlineKeyboardButton(
                f"{label}: {state}",
                callback_data=f"notify_toggle_{key}"
            )
        )
    for index in range(0, len(buttons), 2):
        markup.add(*buttons[index:index + 2])
    markup.add(types.InlineKeyboardButton("Назад", callback_data="notify_back_admin"))
    return markup


pending_actions = {}


@bot.callback_query_handler(func=lambda call: call.data in [
    "change_token", "change_cookies", "change_db_interval",
    "change_balance_interval", "add_admin", "remove_admin",
    "download_db", "show_config", "reload_bot", "notification_settings",
    "migrate_excel_to_db", "generate_all_graphs", "report_day", "report_week"
])
@handler_guard
def callback_admin(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    if call.message.chat.type != 'private':
        return
    if not is_admin(user_id):
        return
    if call.data in ["change_token", "change_cookies", "change_db_interval", "change_balance_interval", "add_admin",
                     "remove_admin"]:
        pending_actions[user_id] = call.data
        field_name = {
            "change_token": "TOKEN",
            "change_cookies": "cookies",
            "change_db_interval": "интервал обновления БД (минуты)",
            "change_balance_interval": "интервал отправки баланса (минуты)",
            "add_admin": "ID нового админа",
            "remove_admin": "ID админа для удаления"
        }[call.data]
        bot.send_message(user_id, f"Отправьте новое значение для: {field_name}")
    elif call.data == "download_db":
        if os.path.exists(DB_FILE):
            bot.send_document(user_id, types.InputFile(DB_FILE))
        else:
            bot.send_message(user_id, MESSAGES['admin_download_not_found'])
    elif call.data == "show_config":
        notification_lines = "\n".join(
            f"{NOTIFICATION_LABELS[key]}: <code>{format_notification_state(value)}</code>"
            for key, value in get_notification_settings().items()
        )
        risk_lines = "\n".join(
            f"{key}: <code>{value}</code>"
            for key, value in get_risk_settings().items()
        )
        api_lines = "\n".join(
            f"{key}: <code>{'***' if key == 'token' and value else value}</code>"
            for key, value in get_api_settings().items()
        )
        bot_status = "Обычный режим"
        conf_text = (
            f"<b>{MESSAGES['config_title']}</b>\n\n"
            f"Состояние: <code>{bot_status}</code>\n"
            f"TOKEN: <code>{config.get('TOKEN', '')}</code>\n"
            f"cookies: <code>{config.get('cookies', '')}</code>\n"
            f"admins: <code>{config.get('admins', [])}</code>\n"
            f"db_update_interval: <code>{config.get('db_update_interval', 30)}</code> минут\n"
            f"balance_send_interval: <code>{config.get('balance_send_interval', 30)}</code> минут\n"
            f"chat_id: <code>{config.get('chat_id', '')}</code>\n\n"
            f"<b>Уведомления</b>\n{notification_lines}"
            f"\n\n<b>Риск-правила</b>\n{risk_lines}"
            f"\n\n<b>Local API</b>\n{api_lines}"
        )
        bot.send_message(user_id, conf_text, parse_mode='HTML')
    elif call.data == "notification_settings":
        bot.send_message(user_id, build_notification_settings_text(), reply_markup=get_notification_settings_panel())
    elif call.data == "reload_bot":
        reload_config(bot)
        bot.send_message(user_id, MESSAGES['admin_reload_success'])
    elif call.data == "migrate_excel_to_db":
        if migrate_excel_to_db():
            bot.send_message(user_id, MESSAGES['migrate_ok'])
        else:
            bot.send_message(user_id, MESSAGES['migrate_fail'])
    elif call.data == "generate_all_graphs":
        count_new = generate_all_graphs()
        bot.send_message(user_id, f"Генерация графиков завершена. Сгенерировано: {count_new} новых графиков.")
    elif call.data == "report_day":
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=1)
        bot.send_message(user_id, build_closed_bots_report("за 24 часа", start_dt, end_dt))
    elif call.data == "report_week":
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=7)
        bot.send_message(user_id, build_closed_bots_report("за 7 дней", start_dt, end_dt))


@bot.message_handler(func=lambda message: message.from_user.id in pending_actions)
@handler_guard
def admin_input_handler(message):
    user_id = message.from_user.id
    action = pending_actions[user_id]
    should_reload = False
    reply_text = None
    try:
        if action == "change_token":
            config['TOKEN'] = message.text.strip()
            should_reload = True
            reply_text = "TOKEN обновлён и применён без перезапуска процесса."
        elif action == "change_cookies":
            config['cookies'] = message.text.strip()
            global WAITING_FOR_RENEW
            WAITING_FOR_RENEW = False
            should_reload = True
            reply_text = "cookies обновлены и будут использованы без перезапуска процесса."
        elif action == "change_db_interval":
            interval = int(message.text.strip())
            config['db_update_interval'] = interval
            should_reload = True
            reply_text = f"Интервал обновления БД теперь {interval} минут."
        elif action == "change_balance_interval":
            interval = int(message.text.strip())
            config['balance_send_interval'] = interval
            should_reload = True
            reply_text = f"Интервал отправки баланса теперь {interval} минут."
        elif action == "add_admin":
            new_admin = int(message.text.strip())
            if new_admin not in config['admins']:
                config['admins'].append(new_admin)
                reply_text = f"Админ {new_admin} добавлен."
            else:
                reply_text = f"{new_admin} уже админ."
        elif action == "remove_admin":
            remove_id = int(message.text.strip())
            if remove_id in config['admins']:
                config['admins'].remove(remove_id)
                reply_text = f"Админ {remove_id} удалён."
            else:
                reply_text = f"{remove_id} не найден в списке админов."

        save_config(config)
        if should_reload:
            reload_config(bot)
        if reply_text:
            bot.send_message(user_id, reply_text)
    except ValueError:
        bot.send_message(user_id, "Некорректный ввод.")
    del pending_actions[user_id]


@bot.callback_query_handler(func=lambda call: call.data == "notify_back_admin" or call.data.startswith("notify_toggle_"))
@handler_guard
def callback_notification_settings(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    if call.message.chat.type != 'private':
        return
    if not is_admin(user_id):
        return

    if call.data == "notify_back_admin":
        try:
            bot.edit_message_text(
                MESSAGES['admin_panel_title'],
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                reply_markup=get_admin_panel()
            )
        except Exception:
            bot.send_message(call.message.chat.id, MESSAGES['admin_panel_title'], reply_markup=get_admin_panel())
        return

    setting_key = call.data.replace("notify_toggle_", "", 1)
    if setting_key not in NOTIFICATION_SETTINGS_DEFAULTS:
        return

    settings = get_notification_settings()
    settings[setting_key] = not settings.get(setting_key, False)
    config["notification_settings"] = settings
    save_config(config)
    bot.edit_message_text(
        build_notification_settings_text(),
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        reply_markup=get_notification_settings_panel()
    )


def apply_runtime_config(config_data):
    global config, TOKEN, cookies, admins, db_update_interval, balance_send_interval, chat_id
    config = apply_config_defaults(config_data)
    TOKEN = config.get('TOKEN', '')
    cookies = config.get('cookies', '')
    admins = config.get('admins', [])
    db_update_interval = config.get('db_update_interval', 30)
    balance_send_interval = config.get('balance_send_interval', 30)
    chat_id = config.get('chat_id', '')
    try:
        if TOKEN:
            bot.token = TOKEN
    except Exception:
        pass


def reload_config(bot_obj=None):
    apply_runtime_config(load_config())
    refresh_api_server()


def update_config_entries(updates):
    current = load_config()
    for key, value in (updates or {}).items():
        if key in {"notification_settings", "risk_settings", "api_settings"} and isinstance(value, dict):
            merged_value = dict(current.get(key) or {})
            merged_value.update(value)
            current[key] = merged_value
        else:
            current[key] = value
    save_config(current)
    apply_runtime_config(current)
    refresh_api_server()
    return current


def sanitize_config_for_output(config_data):
    sanitized = dict(config_data or {})
    for key in ("TOKEN", "cookies", "API_KEY", "API_SECRET"):
        if key in sanitized:
            sanitized[key] = "***"
    return sanitized


def collect_active_bot_records():
    records = []
    for bot_data in fetch_bot_list_data():
        archive_record = build_bot_archive_record(bot_data, is_active=True)
        archive_record["profit_usdt"] = get_top_bot_profit_value(archive_record)
        archive_record["profit_percent"] = get_top_bot_percent_value(archive_record)
        records.append(archive_record)
    return records


def collect_archive_records(limit=100):
    if not USE_DB:
        return []
    ensure_db_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT *
        FROM bot_archive
        ORDER BY COALESCE(ended_ts, created_ts, 0) DESC
        LIMIT ?
        """,
        (int(limit),)
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows


def collect_latest_balance_snapshot():
    if not USE_DB:
        return None
    ensure_db_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM balances ORDER BY date DESC LIMIT 1")
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def execute_readonly_query(sql, params=None):
    query_text = str(sql or "").strip()
    if not query_text:
        raise ValueError("Пустой SQL")
    lowered = query_text.lower()
    if not lowered.startswith("select "):
        raise ValueError("Разрешены только SELECT-запросы")
    if ";" in query_text:
        raise ValueError("Несколько SQL-выражений запрещены")
    if not USE_DB:
        raise ValueError("DB недоступна")
    ensure_db_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(query_text, tuple(params or []))
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows


class LocalApiHandler(BaseHTTPRequestHandler):
    server_version = "tgbybit-local-api/1.0"

    def log_message(self, format_string, *args):
        logging.info("API: " + format_string, *args)

    def _read_json(self):
        content_length = int(self.headers.get("Content-Length", "0") or 0)
        if content_length <= 0:
            return {}
        raw_body = self.rfile.read(content_length)
        return json.loads(raw_body.decode("utf-8"))

    def _send_json(self, status_code, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _authorized(self):
        api_settings = get_api_settings()
        expected_token = str(api_settings.get("token") or "")
        if not expected_token:
            return True
        supplied_token = self.headers.get("X-API-Key") or ""
        return supplied_token == expected_token

    def _route(self):
        parsed = urlparse(self.path)
        return parsed.path, parse_qs(parsed.query)

    def do_GET(self):
        if not self._authorized():
            self._send_json(401, {"ok": False, "error": "unauthorized"})
            return
        path, query = self._route()
        try:
            if path == "/api/health":
                self._send_json(200, {"ok": True, "time": datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
                return
            if path == "/api/config":
                self._send_json(200, {"ok": True, "config": sanitize_config_for_output(config)})
                return
            if path == "/api/balance/latest":
                self._send_json(200, {"ok": True, "balance": collect_latest_balance_snapshot()})
                return
            if path == "/api/bots/active":
                self._send_json(200, {"ok": True, "items": collect_active_bot_records()})
                return
            if path == "/api/bots/archive":
                limit = int((query.get("limit") or ["100"])[0])
                self._send_json(200, {"ok": True, "items": collect_archive_records(limit=limit)})
                return
            if path == "/api/bybit/bots":
                scope = (query.get("scope") or ["active"])[0]
                if scope == "history":
                    data = fetch_historical_bot_list_data()
                else:
                    data = fetch_bot_list_data()
                self._send_json(200, {"ok": True, "scope": scope, "items": data})
                return
            if path == "/api/report/day":
                end_dt = datetime.now()
                start_dt = end_dt - timedelta(days=1)
                self._send_json(200, {"ok": True, "text": build_closed_bots_report("за 24 часа", start_dt, end_dt)})
                return
            if path == "/api/report/week":
                end_dt = datetime.now()
                start_dt = end_dt - timedelta(days=7)
                self._send_json(200, {"ok": True, "text": build_closed_bots_report("за 7 дней", start_dt, end_dt)})
                return
            self._send_json(404, {"ok": False, "error": "not_found"})
        except Exception as e:
            self._send_json(500, {"ok": False, "error": str(e)})

    def do_POST(self):
        if not self._authorized():
            self._send_json(401, {"ok": False, "error": "unauthorized"})
            return
        path, _ = self._route()
        try:
            payload = self._read_json()
            if path == "/api/config":
                allowed_updates = {}
                for key in (
                    "TOKEN", "cookies", "admins", "db_update_interval", "balance_send_interval", "chat_id",
                    "notification_settings", "risk_settings", "api_settings"
                ):
                    if key in payload:
                        allowed_updates[key] = payload[key]
                updated = update_config_entries(allowed_updates)
                self._send_json(200, {"ok": True, "config": sanitize_config_for_output(updated)})
                return
            if path == "/api/db/query":
                rows = execute_readonly_query(payload.get("sql"), payload.get("params"))
                self._send_json(200, {"ok": True, "rows": rows})
                return
            if path == "/api/actions/sync":
                sync_saved = sync_bot_archive(force=True, include_active=True, include_history=True)
                repaired = repair_bot_archive_metrics()
                risk_alerts = dispatch_active_bot_risk_alerts()
                close_alerts = dispatch_bot_close_notifications()
                self._send_json(
                    200,
                    {
                        "ok": True,
                        "sync_saved": sync_saved,
                        "archive_repaired": repaired,
                        "risk_alerts": risk_alerts,
                        "close_alerts": close_alerts
                    }
                )
                return
            self._send_json(404, {"ok": False, "error": "not_found"})
        except Exception as e:
            self._send_json(500, {"ok": False, "error": str(e)})


def start_api_server():
    api_settings = get_api_settings()
    if not api_settings.get("enabled", True):
        return False
    if API_SERVER_STATE["server"] is not None:
        return True
    try:
        server = ThreadingHTTPServer((api_settings.get("host", "127.0.0.1"), int(api_settings.get("port", 8877))), LocalApiHandler)
    except OSError as e:
        logging.error(f"Ошибка запуска local API: {e}")
        return False

    def serve():
        try:
            server.serve_forever()
        except Exception:
            logging.exception("Local API server crashed")

    thread = threading.Thread(target=serve, daemon=True)
    thread.start()
    API_SERVER_STATE["server"] = server
    API_SERVER_STATE["thread"] = thread
    return True


def stop_api_server():
    server = API_SERVER_STATE.get("server")
    thread = API_SERVER_STATE.get("thread")
    if server is None:
        return False
    try:
        server.shutdown()
        server.server_close()
    except Exception:
        logging.exception("Ошибка остановки local API")
    API_SERVER_STATE["server"] = None
    API_SERVER_STATE["thread"] = None
    if thread and thread.is_alive():
        thread.join(timeout=2)
    return True


def refresh_api_server():
    api_settings = get_api_settings()
    server = API_SERVER_STATE.get("server")
    if not api_settings.get("enabled", True):
        stop_api_server()
        return False
    desired_host = api_settings.get("host", "127.0.0.1")
    desired_port = int(api_settings.get("port", 8877))
    if server is not None:
        current_host, current_port = server.server_address[:2]
        if current_host == desired_host and int(current_port) == desired_port:
            return True
        stop_api_server()
    return start_api_server()


# ------------------ ОБРАБОТЧИКИ ДЛЯ ПОЛЬЗОВАТЕЛЬСКОГО МЕНЮ ------------------

@bot.message_handler(commands=['start', 'help'])
@handler_guard
def send_welcome(message):
    bot.send_message(message.chat.id, MESSAGES['start_message'], reply_markup=user_keyboard)


@bot.message_handler(
    func=lambda m: m.text in [MESSAGES['menu_balance'], MESSAGES['menu_graph'], MESSAGES['menu_top'], MESSAGES['menu_admin']])
@handler_guard
def handle_user_menu(message):
    if message.text == MESSAGES['menu_balance']:
        balance_cmd(message)
    elif message.text == MESSAGES['menu_graph']:
        graph_cmd(message)
    elif message.text == MESSAGES['menu_top']:
        top_cmd(message)
    elif message.text == MESSAGES['menu_admin']:
        admin_panel(message)


@bot.message_handler(commands=['balance'])
@handler_guard
def balance_cmd(message):
    try:
        # При команде balance обновляем данные в БД
        balance_info = fetch_balance(add_to_db=True, bot_obj=bot)
        bot.send_message(message.chat.id, balance_info)
    except Exception:
        bot.send_message(message.chat.id, MESSAGES['error_balance'])


def send_admin_panel_message(user_id):
    if not is_admin(user_id):
        return False
    bot.send_message(user_id, MESSAGES['admin_panel_title'], reply_markup=get_admin_panel())
    return True


def show_graph_overview(chat_id, selected_date, message_id=None, viewer_user_id=None):
    bots_data = fetch_bot_list_data()[:6]
    file_path, error = generate_graph_for_date(selected_date, bot_obj=bot, force_refresh=True, bots_data=bots_data)
    if error:
        if message_id is None:
            bot.send_message(chat_id, error)
        else:
            bot.edit_message_caption(chat_id=chat_id, message_id=message_id, caption=error)
        return False

    caption = f"График за {selected_date.strftime('%d_%m_%Y')}\nВыберите дату или карточку бота:"
    markup = build_graph_message_markup(selected_date, bots_data=bots_data, viewer_user_id=viewer_user_id)
    if message_id is None:
        bot.send_photo(chat_id, types.InputFile(file_path), caption=caption, reply_markup=markup)
    else:
        bot.edit_message_media(types.InputMediaPhoto(types.InputFile(file_path)), chat_id=chat_id, message_id=message_id)
        bot.edit_message_caption(chat_id=chat_id, message_id=message_id, caption=caption, reply_markup=markup)
    return True


def show_bot_card(chat_id, message_id, selected_date, bot_index, viewer_user_id=None):
    bots_data = fetch_bot_list_data()[:6]
    file_path, error = generate_bot_card_image(selected_date, bot_index, force_refresh=True, bots_data=bots_data)
    if error:
        bot.edit_message_caption(chat_id=chat_id, message_id=message_id, caption=error)
        return False

    snapshot = build_bot_snapshot(bots_data[bot_index], bot_index)
    caption = build_bot_card_caption(snapshot)
    markup = build_bot_card_markup(
        selected_date,
        bot_index,
        bots_data=bots_data,
        viewer_user_id=viewer_user_id
    )
    bot.edit_message_media(types.InputMediaPhoto(types.InputFile(file_path)), chat_id=chat_id, message_id=message_id)
    bot.edit_message_caption(chat_id=chat_id, message_id=message_id, caption=caption, reply_markup=markup)
    return True


@bot.message_handler(commands=['graph'])
@handler_guard
def graph_cmd(message):
    try:
        # Обновляем данные в БД перед генерацией графика
        fetch_balance(add_to_db=True, bot_obj=bot)
        all_dates = get_all_dates()
        if not all_dates:
            bot.send_message(message.chat.id, "Нет данных для построения графиков.")
            return
        selected_date = all_dates[-1]
        show_graph_overview(message.chat.id, selected_date, viewer_user_id=message.from_user.id)
    except Exception:
        bot.send_message(message.chat.id, MESSAGES['error_graph'])


@bot.message_handler(commands=['top'])
@handler_guard
def top_cmd(message):
    try:
        show_top_bots(message.chat.id, sort_mode="earnings", refresh_archive=True)
    except Exception as e:
        logging.error(f"Ошибка генерации топа ботов: {e}")
        bot.send_message(message.chat.id, "Ошибка генерации топа ботов.")


@bot.message_handler(commands=['migrate_excel'])
@handler_guard
def migrate_excel_command(message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        bot.send_message(message.chat.id, MESSAGES['admin_no_access'])
        return
    if migrate_excel_to_db():
        bot.send_message(message.chat.id, MESSAGES['migrate_ok'])
    else:
        bot.send_message(message.chat.id, MESSAGES['migrate_fail'])


@bot.message_handler(commands=['generate_images'])
@handler_guard
def generate_images_command(message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        bot.send_message(message.chat.id, MESSAGES['admin_no_access'])
        return
    cnt = 0  # Заглушка для генерации картинок
    bot.send_message(message.chat.id, f"{MESSAGES['gen_images_done']} (сгенерировано: {cnt})")


@bot.message_handler(commands=['report_day'])
@handler_guard
def report_day_command(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, MESSAGES['admin_no_access'])
        return
    end_dt = datetime.now()
    start_dt = end_dt - timedelta(days=1)
    bot.send_message(message.chat.id, build_closed_bots_report("за 24 часа", start_dt, end_dt))


@bot.message_handler(commands=['report_week'])
@handler_guard
def report_week_command(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, MESSAGES['admin_no_access'])
        return
    end_dt = datetime.now()
    start_dt = end_dt - timedelta(days=7)
    bot.send_message(message.chat.id, build_closed_bots_report("за 7 дней", start_dt, end_dt))


@bot.callback_query_handler(func=lambda call: call.data.startswith("alert_"))
@handler_guard
def callback_alert(call):
    bot.answer_callback_query(call.id)
    if not is_admin(call.from_user.id):
        return

    if call.data == "alert_mute_30m":
        MARKET_ALERT_STATE["mute_until_ts"] = time.time() + MARKET_ALERT_MUTE_SECONDS
        MARKET_ALERT_STATE["last_sent_minute_key"] = None
        mute_until = datetime.now() + timedelta(seconds=MARKET_ALERT_MUTE_SECONDS)
        try:
            bot.edit_message_reply_markup(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                reply_markup=None
            )
        except Exception:
            pass
        bot.send_message(
            call.message.chat.id,
            f"Сигнализация падения отключена до {mute_until.strftime('%d.%m.%Y %H:%M:%S')}."
        )
        return

    if call.data.startswith("alert_risk_mute_30m_"):
        alert_type = call.data.replace("alert_risk_mute_30m_", "", 1)
        if alert_type not in NOTIFICATION_SETTINGS_DEFAULTS:
            return
        RISK_ALERT_STATE["mute_until_by_type"][alert_type] = time.time() + RISK_ALERT_MUTE_SECONDS
        mute_until = datetime.now() + timedelta(seconds=RISK_ALERT_MUTE_SECONDS)
        try:
            bot.edit_message_reply_markup(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                reply_markup=None
            )
        except Exception:
            pass
        bot.send_message(
            call.message.chat.id,
            f"Риск-уведомления типа «{NOTIFICATION_LABELS.get(alert_type, alert_type)}» отложены до {mute_until.strftime('%d.%m.%Y %H:%M:%S')}."
        )
        return

    if call.data.startswith("alert_risk_disable_"):
        alert_type = call.data.replace("alert_risk_disable_", "", 1)
        if alert_type not in NOTIFICATION_SETTINGS_DEFAULTS:
            return
        update_config_entries({"notification_settings": {alert_type: False}})
        try:
            bot.edit_message_reply_markup(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                reply_markup=None
            )
        except Exception:
            pass
        bot.send_message(
            call.message.chat.id,
            f"Уведомления типа «{NOTIFICATION_LABELS.get(alert_type, alert_type)}» отключены."
        )


@bot.callback_query_handler(func=lambda call: call.data.startswith("top_"))
@handler_guard
def callback_top(call):
    bot.answer_callback_query(call.id)
    if call.data == "top_noop":
        return
    sort_mode = "earnings"
    page = 0
    if call.data.startswith("top_view_"):
        payload = call.data.replace("top_view_", "", 1)
        mode_part, _, page_part = payload.rpartition("_")
        if mode_part:
            sort_mode = mode_part
        page = safe_int(page_part) or 0
    elif call.data.startswith("top_mode_"):
        sort_mode = call.data.replace("top_mode_", "", 1)
    show_top_bots(
        call.message.chat.id,
        sort_mode=sort_mode,
        page=page,
        message_id=call.message.message_id,
        refresh_archive=False
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("graph_"))
@handler_guard
def callback_graph(call):
    bot.answer_callback_query(call.id)
    data_str = call.data
    if data_str == "graph_admin_back":
        if send_admin_panel_message(call.from_user.id):
            pass
        else:
            bot.send_message(call.message.chat.id, MESSAGES['admin_no_access'])
        return
    if data_str.startswith("graph_overview_"):
        date_token = data_str.replace("graph_overview_", "")
        selected_date = token_to_date(date_token)
        show_graph_overview(
            call.message.chat.id,
            selected_date,
            message_id=call.message.message_id,
            viewer_user_id=call.from_user.id
        )
        return
    if data_str.startswith("graph_bot_"):
        _, _, date_token, bot_index = data_str.split("_")
        selected_date = token_to_date(date_token)
        show_bot_card(
            call.message.chat.id,
            call.message.message_id,
            selected_date,
            int(bot_index),
            viewer_user_id=call.from_user.id
        )
        return
    if data_str.startswith("graph_day_"):
        date_str = data_str.replace("graph_day_", "")
        selected_date = datetime.strptime(date_str, "%d_%m_%Y").date()
        show_graph_overview(
            call.message.chat.id,
            selected_date,
            message_id=call.message.message_id,
            viewer_user_id=call.from_user.id
        )
    elif data_str.startswith("graph_month_"):
        ym_str = data_str.replace("graph_month_", "")
        parts = ym_str.split("_")
        if len(parts) != 2:
            return
        year = int(parts[0])
        month = int(parts[1])
        d_list = get_all_dates()
        md = dates_in_month(d_list, year, month)
        if not md:
            return
        selected_date = md[-1]
        show_graph_overview(
            call.message.chat.id,
            selected_date,
            message_id=call.message.message_id,
            viewer_user_id=call.from_user.id
        )
    elif data_str.startswith("graph_monthnav_prev_") or data_str.startswith("graph_monthnav_next_"):
        ym_str = data_str.split("_")[-2] + "_" + data_str.split("_")[-1]
        parts = ym_str.split("_")
        if len(parts) != 2:
            return
        year = int(parts[0])
        month = int(parts[1])
        d_list = get_all_dates()
        all_months = get_months_from_dates(d_list)
        if (year, month) not in all_months:
            return
        idx = all_months.index((year, month))
        if data_str.startswith("graph_monthnav_prev_") and idx > 0:
            year, month = all_months[idx - 1]
        elif data_str.startswith("graph_monthnav_next_") and idx < len(all_months) - 1:
            year, month = all_months[idx + 1]
        md = dates_in_month(d_list, year, month)
        if not md:
            return
        selected_date = md[-1]
        show_graph_overview(
            call.message.chat.id,
            selected_date,
            message_id=call.message.message_id,
            viewer_user_id=call.from_user.id
        )


def run_bot_polling():
    while True:
        try:
            bot.infinity_polling(
                timeout=30,
                long_polling_timeout=30,
                logger_level=logging.ERROR
            )
        except KeyboardInterrupt:
            raise
        except ApiTelegramException as e:
            if getattr(e, "error_code", None) == 409 or "terminated by other getUpdates request" in str(e):
                logging.error("Polling stopped: another bot instance is already using getUpdates.")
                stop_all_threads()
                return
            logging.error(f"Ошибка polling Telegram: {e}")
            sleep(5)
        except Exception as e:
            logging.error(f"Ошибка polling Telegram: {e}")
            sleep(5)


if __name__ == '__main__':
    lock_acquired, owner_pid = acquire_instance_lock()
    if not lock_acquired:
        logging.error(
            "Another tgbybit.py instance is already running%s. Exiting.",
            f" (PID {owner_pid})" if owner_pid else ""
        )
        raise SystemExit(1)
    if USE_DB:
        create_db()
        ensure_db_schema()
        repair_balance_history()
        repair_duplicate_bot_balance_spikes(limit_rows=None)
        repair_bot_archive_metrics()
        bootstrap_bot_close_notifications()
    start_api_server()
    start_threads()
    run_bot_polling()
